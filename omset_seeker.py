"""
OMSET SEEKER
Query cepat omset (KRT) per outlet, lintas semua brand, dari hasil transpose CSV.

Sumber data: omset_pipeline/output/CSV/{UMUM,HOREKA}/OMSHAR {tipe} {brand} TRANSPOSED_{sheet}_query.csv
(dihasilkan oleh transpose.py — header 8 baris sudah di-strip di file *_query.csv)
"""

import os
import re
import shutil
from functools import lru_cache
from pathlib import Path

import openpyxl
import pandas as pd

CSV_DIR = Path(__file__).resolve().parent / "omset_pipeline" / "output" / "CSV"
CACHE_DIR = CSV_DIR.parent / "CACHE"
TOKO_GABUNGAN_DIR = Path(r"D:\Data BIA\INFO BIA\Toko Gabungan")

# UMUM: grup gabungan sumbernya file Excel bulanan dari divisi lain, satu file
# berisi banyak grup (lihat find_latest_toko_gabungan()). HOREKA tidak punya
# proses/file semacam itu sama sekali -- jadi pakai file TERPISAH di folder yang
# sama, format beda (satu sheet = satu grup), lihat find_latest_horeka_gabungan().

COL_WIL, COL_SITE, COL_CUST = 0, 1, 2
COL_PROPINSI, COL_KOTA, COL_KECAMATAN, COL_ALAMAT = 11, 12, 13, 18
OMSET_COLS = list(range(140, 164))  # EK:FH = JAN 2025 - DES 2026
EXTRA_COLS = [COL_PROPINSI, COL_KOTA, COL_KECAMATAN, COL_ALAMAT]
EXTRA_NAMES = ["Propinsi", "Kota", "Kecamatan", "Alamat"]

MONTH_LABELS = [
    f"{m} {y}"
    for y in ["2025", "2026"]
    for m in ["JAN", "FEB", "MAR", "APR", "MEI", "JUN", "JUL", "AGS", "SEP", "OKT", "NOV", "DES"]
]

# 20 brand SKU (sama dengan transpose.py)
BRAND_ORDER = [
    "ABIDIN", "AMERAJA", "APIDIN", "SOMAEK",
    "SIMER", "SIJO", "SIDU", "SIRAK",
    "SPA FILTERED", "SPA UNFILTERED", "SINGARAJA",
    "PROST PILSENER", "PROST LAGER", "PRL LAGER",
    "PRL APPLE LIME", "PRL RASPBERRY",
    "PROST ALSTER", "KALTENBERG", "KONIG WEISSBIER", "KONIG DUNKEL",
]

# KONIG WEISSBIER tidak murni dari file WBR -- WBR berisi Weissbier+Dunkel
# gabungan. OMSET asli (VLOOKUP formula) mengurangi KONIG DUNKEL dari WBR
# untuk dapat angka Weissbier murni: lihat seek_outlet().
KONIG_WEISSBIER_RAW = "KONIG WEISSBIER"
KONIG_DUNKEL = "KONIG DUNKEL"

# Varian "WITH KEG" -- HOREKA-only, dipakai outlet yang jual bir keg/draft
# (lihat [[project-omshar-technical]] / sheet "WITH KEG" di OMSET HOREKA 2026.xlsx).
# Bukan bagian dari BRAND_ORDER default -- panggil seek_outlet(..., brands=
# BRAND_ORDER + HOREKA_KEG_BRAND_ORDER) atau render_outlet_report(with_keg=True)
# untuk outlet yang butuh baris ini.
HOREKA_KEG_BRAND_ORDER = [
    "SINGARAJA KEG 10", "SINGARAJA KEG 20", "SINGARAJA KEG 30",
    "SINGARAJA BREMER HWG",
    "PROST PILSENER KEG 10", "PROST PILSENER KEG 20", "PROST PILSENER KEG 30",
    "PROST LAGER KEG 10", "PROST LAGER KEG 20", "PROST LAGER KEG 30",
    "PROST RAJAWALI KEG 10", "PROST RAJAWALI KEG 30",
    "SINGARAJA PET 20L", "PROST PILSENER PET 20L", "PROST LAGER PET 20L", "PROST RAJAWALI PET 20L",
]

# "BIR" = total kategori (bukan SKU individu), "DIV AB1" = total divisi
# (lebih luas dari BIR). Keduanya punya file/transpose sendiri, dipakai untuk
# cross-check + baris ringkasan di laporan (lihat get_brand_months/compute_bev).
VALIDATION_BRAND = "BIR"
DIV_AB1_BRAND = "DIV AB1"


def resolve_site_list(site: str, omshar_type: str = "UMUM") -> list[str]:
    """Kalau `site` adalah kode toko gabungan (lihat load_gabungan_map), kembalikan daftar
    kode site toko anaknya (SEMUA channel dicampur jadi satu list -- dipakai pemanggil
    yang cuma query SATU channel, mis. sku_lookup.py; anggota dari channel lain di situ
    otomatis tidak match apa pun, bukan salah, cuma tidak ikut lintas-channel di jalur
    itu). Kalau bukan gabungan, kembalikan [site] apa adanya."""
    gabungan = load_gabungan_map(omshar_type).get(site)
    return gabungan["children"] if gabungan else [site]


def resolve_site_list_by_channel(site: str, omshar_type: str = "UMUM") -> dict:
    """Return {channel: [site_codes]} -- biasanya cuma {omshar_type: [site]} (site
    tunggal) atau {omshar_type: [semua anak]} (gabungan biasa, satu channel).

    HOREKA Gabungan mendukung PENGECUALIAN per-outlet lewat kolom CHANNEL opsional
    di file-nya: kalau satu outlet anggota grup ternyata cuma terdaftar di UMUM
    (bukan salah ketik, memang begitu adanya di data OMSHAR), ditandai CHANNEL=UMUM
    di baris itu supaya datanya tetap ikut dijumlah dari channel yang benar -- tanpa
    ini, outlet itu akan selalu nol di laporan HOREKA (silently), karena dia memang
    tidak punya baris apa pun di data HOREKA sama sekali."""
    gabungan = load_gabungan_map(omshar_type).get(site)
    if not gabungan:
        return {omshar_type: [site]}
    by_channel = gabungan.get("children_by_channel")
    if by_channel:
        return {ch: sites for ch, sites in by_channel.items() if sites}
    return {omshar_type: gabungan["children"]}


def get_brand_months(brand: str, site: str, omshar_type: str = "UMUM") -> dict:
    """Ambil nilai KRT per bulan untuk satu brand + site (0 semua jika tidak ada).
    Kalau `site` adalah kode toko gabungan, otomatis dijumlah dari semua toko anaknya
    -- LINTAS CHANNEL kalau ada pengecualian (lihat resolve_site_list_by_channel())."""
    total = {m: 0.0 for m in MONTH_LABELS}
    for channel, sites in resolve_site_list_by_channel(site, omshar_type).items():
        match = query_brand(brand, sites, channel)
        if not match.empty:
            summed = match[MONTH_LABELS].fillna(0).sum().to_dict()
            for m in MONTH_LABELS:
                total[m] += summed[m]
    return total


def compute_bev(site: str, omshar_type: str = "UMUM") -> dict:
    """BEV tidak punya file sendiri -- dihitung sebagai DIV AB1 - BIR per bulan."""
    div_vals = get_brand_months(DIV_AB1_BRAND, site, omshar_type)
    bir_vals = get_brand_months(VALIDATION_BRAND, site, omshar_type)
    return {m: div_vals[m] - bir_vals[m] for m in MONTH_LABELS}


def _csv_path(omshar_type: str, brand: str, sheet: str) -> Path:
    subdir = "UMUM" if omshar_type == "UMUM" else "HOREKA"
    fname = f"OMSHAR {omshar_type} {brand} TRANSPOSED_{sheet}_query.csv"
    return CSV_DIR / subdir / fname


def _read_query_csv(path: Path) -> pd.DataFrame:
    cols_needed = [COL_WIL, COL_SITE, COL_CUST] + EXTRA_COLS + OMSET_COLS
    df = pd.read_csv(
        path,
        header=None,
        usecols=cols_needed,
        dtype={COL_SITE: str},
        encoding="utf-8-sig",
    )
    df.columns = ["Wilayah", "Site", "Outlet"] + EXTRA_NAMES + MONTH_LABELS
    df["Site"] = df["Site"].str.strip()
    return df


@lru_cache(maxsize=None)
def get_cutoff_parts(brand: str = VALIDATION_BRAND, omshar_type: str = "UMUM", sheet: str | None = None):
    """Ambil (day, month, year) cut off dari baris 'PERIODE : JAN sd DD/MM/YYYY', atau None kalau gagal parse.

    Di-cache karena sekarang dipanggil PER BRAND (bukan cuma sekali per outlet
    pakai VALIDATION_BRAND) buat kolom Cut Off per baris di tabel web -- tanpa
    cache, satu pencarian outlet bisa baca ~20-29 file CSV kecil berulang kali.
    Sama seperti load_brand()/load_gabungan_map(), HARUS di-cache_clear() setelah
    Transpose supaya tidak diam-diam nyangkut di cutoff lama."""
    sheet = sheet or ("DAPUL" if omshar_type == "UMUM" else "HOREKA")
    full_path = _csv_path(omshar_type, brand, sheet).with_name(
        _csv_path(omshar_type, brand, sheet).name.replace("_query", "")
    )
    if not full_path.exists():
        return None
    periode_row = pd.read_csv(full_path, header=None, nrows=3, encoding="utf-8-sig").iloc[2, 0]
    date_part = str(periode_row).split("sd")[-1].strip()
    try:
        day, month, year = date_part.split("/")
        return int(day), int(month), int(year)
    except ValueError:
        return None


def get_cutoff_date(brand: str = VALIDATION_BRAND, omshar_type: str = "UMUM", sheet: str | None = None) -> str:
    """Ambil tanggal cut off dari baris 'PERIODE : JAN sd DD/MM/YYYY' di file CSV penuh (bukan _query)."""
    parts = get_cutoff_parts(brand, omshar_type, sheet)
    if parts is None:
        return ""
    day, month, year = parts
    months_id = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Ags", "Sep", "Okt", "Nov", "Des"]
    return f"{day} {months_id[month - 1]} {year}"


def _source_csv_paths(brand: str, omshar_type: str) -> list[Path]:
    sheets = ("DAPUL", "LAPUL") if omshar_type == "UMUM" else ("HOREKA",)
    return [p for p in (_csv_path(omshar_type, brand, s) for s in sheets) if p.exists()]


def _parquet_path(brand: str, omshar_type: str) -> Path:
    subdir = "UMUM" if omshar_type == "UMUM" else "HOREKA"
    return CACHE_DIR / subdir / f"OMSHAR {omshar_type} {brand}.parquet"


@lru_cache(maxsize=None)
def load_brand(brand: str, omshar_type: str = "UMUM") -> pd.DataFrame:
    """Load + cache semua sheet wilayah untuk satu brand (DAPUL+LAPUL utk UMUM, HOREKA utk HOREKA).

    Disk cache: hasil parse CSV disimpan sebagai parquet di output/CACHE/ -- baca parquet
    ~30x lebih cepat daripada parse CSV (~1 GB per grup), jadi pencarian pertama setelah
    app restart tidak lagi ~20 detik. Invalidasi OTOMATIS: kalau ada CSV sumber yang LEBIH
    BARU dari parquet-nya (mis. habis transpose, termasuk lewat RUN.bat manual), parse
    ulang -- ditambah clear eksplisit lewat clear_brand_cache() dari halaman Transpose."""
    sources = _source_csv_paths(brand, omshar_type)
    if not sources:
        return pd.DataFrame(columns=["Wilayah", "Site", "Outlet"] + MONTH_LABELS)

    pq_path = _parquet_path(brand, omshar_type)
    if pq_path.exists():
        newest_csv = max(p.stat().st_mtime for p in sources)
        if pq_path.stat().st_mtime >= newest_csv:
            try:
                return pd.read_parquet(pq_path)
            except Exception:
                pass  # cache korup -- jatuh ke parse ulang di bawah

    df = pd.concat([_read_query_csv(p) for p in sources], ignore_index=True)

    try:
        pq_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = pq_path.with_name(f".{pq_path.stem}.{os.getpid()}.tmp.parquet")
        try:
            df.to_parquet(tmp)
            os.replace(tmp, pq_path)
        finally:
            tmp.unlink(missing_ok=True)
    except Exception:
        pass  # gagal tulis cache tidak fatal -- berikutnya cukup parse ulang lagi
    return df


def clear_brand_cache() -> None:
    """Hapus parquet disk cache + lru_cache in-memory load_brand(). Dipanggil setelah
    Transpose (bareng panggilan cache_clear lainnya di pages/1_Sync_dan_Transpose.py).
    mtime-check di load_brand() sebenarnya sudah cukup, ini supaya cache langsung bersih."""
    load_brand.cache_clear()
    shutil.rmtree(CACHE_DIR, ignore_errors=True)


def query_brand(brand: str, site_list: list[str], omshar_type: str = "UMUM") -> pd.DataFrame:
    """Cari baris outlet tertentu di dalam data satu brand."""
    df = load_brand(brand, omshar_type)
    return df[df["Site"].isin(site_list)]


# Satu fungsi query per brand-key, mis. BRAND_LOADERS["ABIDIN"](["0815-..."])
BRAND_LOADERS = {
    brand: (lambda site_list, b=brand, t="UMUM": query_brand(b, site_list, t))
    for brand in BRAND_ORDER
}


_INDO_MONTHS = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
}
_TOKO_GABUNGAN_NAME_RE = re.compile(
    r"^Toko Gabungan Update (\d{1,2}) (\w+) (\d{4})\.xlsx$", re.IGNORECASE
)


def find_latest_toko_gabungan() -> Path | None:
    """File 'Toko Gabungan Update {tgl} {bulan} {tahun}.xlsx' diupdate manual tiap bulan.

    Tanggal diambil dari NAMA FILE (bukan LastWriteTime filesystem) -- LastWriteTime tidak
    bisa dipercaya karena berubah kalau ada file lain di folder itu yang dibuka/disimpan
    orang di Excel (pernah kejadian: file 'Update 1 April 2026' yang lagi dibuka orang
    kelihatan lebih baru dari 'Update 2 Juli 2026' yang isinya benar-benar lebih baru,
    padahal ukurannya 690MB dan bikin lama sekali/macet kalau ke-parse tanpa sengaja).
    File yang namanya tidak persis cocok pola ini (varian '-val', '- Copy', 'REV', dst)
    sengaja diabaikan -- hanya rilis bulanan asli yang dipakai."""
    if not TOKO_GABUNGAN_DIR.exists():
        return None
    dated = []
    for p in TOKO_GABUNGAN_DIR.glob("Toko Gabungan Update*.xlsx"):
        if p.name.startswith("~$"):
            continue
        m = _TOKO_GABUNGAN_NAME_RE.match(p.name)
        if not m:
            continue
        day, month_name, year = m.groups()
        month = _INDO_MONTHS.get(month_name.lower())
        if month is None:
            continue
        dated.append(((int(year), month, int(day)), p))
    if not dated:
        return None
    return max(dated, key=lambda x: x[0])[1]


def _gabungan_from_flat_sheet(ws) -> dict:
    """Format baru (mis. sheet 'Sheet1' di file Juli 2026+): flat, satu baris per toko --
    header WIL/SITE/CUST/GROUP, nama grup langsung di kolom GROUP. Jauh lebih robust
    daripada format block lama karena tidak bergantung pada baris kosong sebagai pemisah."""
    groups: dict[str, dict] = {}  # group name -> {"wilayah":..., "sites": [ordered]}
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header (WIL, SITE, CUST, GROUP)
    for row in rows:
        if len(row) < 4:
            continue
        wilayah, site, _cust, group_name = row[0], row[1], row[2], row[3]
        if not site or not group_name:
            continue
        g = groups.setdefault(str(group_name).strip(), {"wilayah": wilayah, "sites": []})
        g["sites"].append(str(site).strip())

    gabungan_map = {}
    for name, g in groups.items():
        if len(g["sites"]) < 2:
            continue
        # kode gabungan sintetis = site toko pertama (urutan file) + "1" -- diverifikasi
        # cocok persis dengan kode di sheet block lama (mis. A11) untuk grup yang sama.
        gab_site = f"{g['sites'][0]}1"
        gabungan_map[gab_site] = {"name": name, "wilayah": g["wilayah"], "children": g["sites"]}
    return gabungan_map


def _gabungan_from_block_sheet(ws) -> dict:
    """Format lama (mis. file sampai dengan Juni 2026): satu sheet, blok dipisah baris
    kosong. Tiap blok = beberapa baris toko individual (kolom Wilayah/Area kosong) diikuti
    SATU baris ringkasan gabungan (kolom Wilayah/Area terisi -- penanda paling andal,
    bukan nama yang mengandung kata 'GABUNGAN' karena ada variasi penulisan seperti
    'GAB ...' atau '... Grup'). Kode site baris ringkasan = kode site toko pertama dalam
    blok + '1' di akhir."""
    gabungan_map = {}
    block: list[tuple] = []

    def flush(block_rows):
        if len(block_rows) < 2:
            return
        summary = block_rows[-1]
        children = block_rows[:-1]
        gab_site = summary[2]
        if not gab_site:
            return
        gabungan_map[str(gab_site).strip()] = {
            "name": summary[3],
            "wilayah": summary[0],
            "children": [str(r[2]).strip() for r in children if r[2]],
        }

    for row in ws.iter_rows(values_only=True):
        is_blank = all(v is None for v in row)
        if is_blank:
            flush(block)
            block = []
            continue
        block.append(row)
        if row[0] is not None:
            flush(block)
            block = []
    flush(block)

    return gabungan_map


_HOREKA_GABUNGAN_NAME_RE = re.compile(
    r"^Gabungan HOREKA Update (\d{1,2}) (\w+) (\d{4})\.xlsx$", re.IGNORECASE
)


def find_latest_horeka_gabungan() -> Path | None:
    """File 'Gabungan HOREKA Update {tgl} {bulan} {tahun}.xlsx' -- sama folder dan konvensi
    penamaan/pemilihan-versi-terbaru dengan Toko Gabungan UMUM (lihat find_latest_toko_gabungan()
    untuk alasan kenapa tanggal diambil dari NAMA FILE, bukan LastWriteTime), tapi prefix nama
    beda sengaja supaya tidak pernah ke-mix sama file UMUM di folder yang sama."""
    if not TOKO_GABUNGAN_DIR.exists():
        return None
    dated = []
    for p in TOKO_GABUNGAN_DIR.glob("Gabungan HOREKA Update*.xlsx"):
        if p.name.startswith("~$"):
            continue
        m = _HOREKA_GABUNGAN_NAME_RE.match(p.name)
        if not m:
            continue
        day, month_name, year = m.groups()
        month = _INDO_MONTHS.get(month_name.lower())
        if month is None:
            continue
        dated.append(((int(year), month, int(day)), p))
    if not dated:
        return None
    return max(dated, key=lambda x: x[0])[1]


def _gabungan_from_horeka_workbook(wb) -> dict:
    """Format HOREKA: SATU SHEET = SATU GRUP (nama sheet = nama grup), beda dari UMUM yang
    satu file/sheet berisi banyak grup sekaligus. Tiap baris dalam sheet = satu outlet
    anggota grup itu, kolom dicari lewat header baris pertama (SITE/OUTLET/WILAYAH, urutan
    bebas, case-insensitive) -- bukan posisi kolom tetap, supaya tidak gampang salah kalau
    urutan kolom di file beda-beda antar sheet."""
    gabungan_map = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip().upper() if h is not None else "" for h in rows[0]]
        if "SITE" not in header or "WILAYAH" not in header:
            continue  # sheet ini bukan format yang diharapkan, skip diam-diam
        site_idx = header.index("SITE")
        wilayah_idx = header.index("WILAYAH")

        # Kolom CHANNEL opsional -- pengecualian per-outlet: anggota grup yang
        # SEBENARNYA cuma terdaftar di channel lain (mis. UMUM), bukan HOREKA
        # (bukan salah ketik, memang begitu adanya di data OMSHAR mentah).
        # Tanpa ini outlet itu akan selalu 0 di laporan HOREKA karena dia
        # memang tidak punya baris apa pun di data HOREKA sama sekali.
        channel_idx = header.index("CHANNEL") if "CHANNEL" in header else None

        sites, wilayah = [], None
        children_by_channel: dict[str, list[str]] = {}
        for row in rows[1:]:
            if site_idx >= len(row) or not row[site_idx]:
                continue
            site_code = str(row[site_idx]).strip()
            sites.append(site_code)
            if wilayah is None and wilayah_idx < len(row) and row[wilayah_idx]:
                wilayah = str(row[wilayah_idx]).strip()
            channel = "HOREKA"
            if channel_idx is not None and channel_idx < len(row) and row[channel_idx]:
                override = str(row[channel_idx]).strip().upper()
                if override in ("UMUM", "HOREKA"):
                    channel = override
            children_by_channel.setdefault(channel, []).append(site_code)
        if len(sites) < 2:
            continue

        group_name = ws.title.strip()
        # Kode sintetis = toko pertama + angka -- BUKAN selalu '1' seperti UMUM (yang
        # formatnya warisan VLOOKUP asli, jangan diubah supaya tetap cocok kalau ada
        # sistem lain yang bergantung padanya). HOREKA murni bikinan sendiri, jadi bebas
        # dibikin lebih aman: kalau dua grup beda kebetulan sama-sama diawali toko yang
        # sama (kejadian nyata: sheet gabungan besar + sheet pecahannya per wilayah,
        # sama-sama mulai dari toko yang sama), tanpa penomoran ini grup KEDUA akan diam-
        # diam menimpa yang pertama di dict (kunci sama) -- salah satu grup lenyap tanpa
        # pesan error apa pun.
        suffix = 1
        gab_site = f"{sites[0]}{suffix}"
        while gab_site in gabungan_map:
            suffix += 1
            gab_site = f"{sites[0]}{suffix}"
        gabungan_map[gab_site] = {
            "name": group_name,
            "wilayah": wilayah or "-",
            "children": sites,
            "children_by_channel": children_by_channel,
        }
    return gabungan_map


@lru_cache(maxsize=1)
def load_horeka_gabungan_map() -> dict:
    """Grup gabungan HOREKA -- file 'Gabungan HOREKA Update ...' terpisah dari file Toko
    Gabungan UMUM (beda konvensi: satu sheet = satu grup, bukan satu sheet banyak grup),
    karena HOREKA tidak punya proses/file bawaan dari divisi lain sama sekali."""
    path = find_latest_horeka_gabungan()
    if path is None:
        return {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return _gabungan_from_horeka_workbook(wb)


@lru_cache(maxsize=None)
def load_gabungan_map(omshar_type: str = "UMUM") -> dict:
    """Parse grup gabungan jadi {kode_site_gabungan: {name, wilayah, children: [site,...]}}.
    Kode gabungan sintetis (toko pertama + '1') tidak pernah muncul di data OMSHAR mentah.

    UMUM: file Toko Gabungan Excel bulanan dari divisi lain -- format file berubah antar
    bulan, pakai sheet 'Sheet1' (flat, baru) kalau ada, fallback ke format block lama
    (sheet pertama, mis. 'A11') kalau tidak.
    HOREKA: file Gabungan HOREKA terpisah, satu sheet per grup -- lihat load_horeka_gabungan_map()."""
    if omshar_type == "HOREKA":
        return load_horeka_gabungan_map()

    path = find_latest_toko_gabungan()
    if path is None:
        return {}

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Sheet1" in wb.sheetnames:
        return _gabungan_from_flat_sheet(wb["Sheet1"])
    return _gabungan_from_block_sheet(wb.worksheets[0])


def seek_outlet(site: str, omshar_type: str = "UMUM", brands: list[str] | None = None):
    """Bangun tabel OMSET-style (baris=brand, kolom=bulan) untuk satu outlet.

    Kalau `site` cocok dengan kode toko gabungan (lihat load_gabungan_map), data dijumlah
    dari semua toko anak-nya -- kode gabungan itu sendiri tidak pernah ada di data OMSHAR
    mentah, murni identitas turunan dari file Toko Gabungan."""
    brands = brands or BRAND_ORDER

    gabungan = load_gabungan_map(omshar_type).get(site)
    channel_sites = resolve_site_list_by_channel(site, omshar_type)

    rows = []
    outlet_info = None

    for brand in brands:
        row = {m: 0.0 for m in MONTH_LABELS}
        for channel, sites in channel_sites.items():
            match = query_brand(brand, sites, channel)
            if match.empty:
                continue
            summed = match[MONTH_LABELS].fillna(0).sum().to_dict()
            for m in MONTH_LABELS:
                row[m] += summed[m]
            if outlet_info is None:
                outlet_info = match.iloc[0][["Wilayah", "Outlet"] + EXTRA_NAMES].to_dict()
                outlet_info["Site"] = site
        if brand == KONIG_WEISSBIER_RAW:
            dunkel = get_brand_months(KONIG_DUNKEL, site, omshar_type)
            row = {m: row[m] - dunkel[m] for m in MONTH_LABELS}
        row["Brand"] = brand
        rows.append(row)

    table = pd.DataFrame(rows).set_index("Brand")[MONTH_LABELS]

    if gabungan and outlet_info is not None:
        # nama & wilayah pakai identitas gabungan, bukan toko anak pertama yang kebetulan cocok
        outlet_info["Outlet"] = gabungan["name"]
        outlet_info["Wilayah"] = gabungan["wilayah"]

    return table, outlet_info


def build_outlet_index(omshar_type: str) -> pd.DataFrame:
    """Daftar unik (Site, Outlet, Wilayah) per grup -- dipakai buat browse/search
    outlet di berbagai halaman Streamlit. Diambil dari brand BIR (ada di hampir semua
    outlet). Pemanggil disarankan bungkus dengan @st.cache_data karena UMUM sendiri
    ~72rb baris, HOREKA ~15rb -- tidak mau di-dedup ulang tiap keystroke.

    Kode toko gabungan (UMUM: file Toko Gabungan eksternal; HOREKA: didefinisikan manual
    lewat Atur Gabungan HOREKA) ikut disisipkan supaya bisa dicari/dipilih juga -- kode
    ini sendiri tidak ada di data OMSHAR mentah."""
    df = load_brand("BIR", omshar_type)
    idx = df[["Site", "Outlet", "Wilayah"]].drop_duplicates(subset="Site").reset_index(drop=True)

    gabungan = load_gabungan_map(omshar_type)
    if gabungan:
        gab_rows = pd.DataFrame([
            {"Site": site, "Outlet": f"{g['name']} (Gabungan)", "Wilayah": g["wilayah"]}
            for site, g in gabungan.items()
        ])
        idx = pd.concat([gab_rows, idx], ignore_index=True)

    return idx


def display_outlet_report(site: str, omshar_type: str = "UMUM"):
    table, info = seek_outlet(site, omshar_type)
    if info is None:
        print(f"Site '{site}' tidak ditemukan di {omshar_type}.")
        return

    print(f"\n{'=' * 70}")
    print(f"Site    : {site}")
    print(f"Wilayah : {info['Wilayah']}")
    print(f"Outlet  : {info['Outlet']}")
    print(f"{'=' * 70}")

    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 300)
    print(table)

    total_per_brand = table.sum(axis=1)
    grand_total = total_per_brand.sum()
    print("\nTotal per brand (Jan25-Des26), hanya yang ada penjualan:")
    print(total_per_brand[total_per_brand > 0])
    print(f"\nGRAND TOTAL semua brand: {grand_total:,.0f} KRT")

    bir_match = query_brand(VALIDATION_BRAND, [site], omshar_type)
    if not bir_match.empty:
        bir_total = bir_match[MONTH_LABELS].fillna(0).sum().sum()
        print(f"Validasi kategori BIR  : {bir_total:,.0f} KRT  (selisih: {grand_total - bir_total:,.0f})")


if __name__ == "__main__":
    while True:
        site = input("\nSite number (atau 'exit'): ").strip()
        if site.lower() == "exit":
            break
        group = input("Grup [UMUM/HOREKA, default UMUM]: ").strip().upper() or "UMUM"
        display_outlet_report(site, group)
