"""
RENDER BPR
Tampilan (HTML web + export Excel/PDF) untuk Rekap Per Depo / Per Wilayah --
format di-copy PERSIS dari "BPR BIA DAILY TEMPLATE.xlsx" (dicek langsung
lewat openpyxl cell.fill/font/number_format/merged_cells, bukan tebakan):
  - Warna: DATA (NORM/STOK/DOI) kuning FFFF00, PROPOSED ORDER hijau FF00FF00
    (PROST LAGER..WEISSBIER) / oranye FFC000 (SINGARAJA ARAK, BAESOMAEK,
    SINGARAJA PALE ALE), TOTAL merah FF0000, ACTUAL ORDER/% gelap teks putih.
  - Header 2 baris digabung (merged cell): baris atas nama grup (DATA,
    PROPOSED ORDER), baris bawah nama kolom (NORM/STOK/DOI, brand, TOTAL) --
    WILAYAH/DEPO/ACTUAL ORDER/% span vertikal 2 baris tanpa sub-header.
  - Font Calibri BOLD di SEMUA sel (bukan cuma header) -- pilihan gaya
    template asli, dicek langsung (cell.font.bold=True bahkan di baris data).
  - Format angka accounting 0 desimal (dash '-' untuk nol) di SEMUA kolom
    angka termasuk STOK & DOI (bukan 2/1 desimal kayak biasanya di app ini) --
    persis number_format template (`_(* #,##0_);...;_(* "-"??_);...`).
    Kolom % pakai '0%' (0 desimal, bukan 1).
  - Judul besar "BPR BIA" 28pt + subjudul 18pt (diisi tanggal file sumber
    yang SEBENARNYA, bukan label statis "UPDATE ..." yang di template asli
    ternyata sering lupa di-update manual).
"""

import io
import re

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.collections import PatchCollection
    from matplotlib.figure import Figure
    from matplotlib.patches import Rectangle
    MATPLOTLIB_AVAILABLE = True
except Exception as e:  # pragma: no cover
    MATPLOTLIB_AVAILABLE = False
    _MATPLOTLIB_IMPORT_ERROR = e

import bpr_pipeline as bp

C_DATA = "FFFF00"
C_BRAND_1 = "00FF00"
C_BRAND_2 = "FFC000"
C_TOTAL = "FF0000"
C_DARK = "1F1F1F"
C_WHITE = "FFFFFF"

BRAND_GROUP_1 = bp.BRAND_COLUMNS[:7]
BRAND_GROUP_2 = bp.BRAND_COLUMNS[7:]

FONT_NAME = "Calibri, sans-serif"

_RAW_TS_RE = re.compile(r"BPR_BIA-(\d{14})\.")
_MONTH_ID = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
             "Juli", "Agustus", "September", "Oktober", "November", "Desember"]


def format_update_label(source_name: str) -> str:
    """'BPR_BIA-20260824070100.xls' -> 'UPDATE 24 Agustus 2026 07:01' -- dinamis
    dari timestamp NYATA di nama file, gantiin label statis 'UPDATE 18 MEI 2026'
    di template yang sering lupa di-update manual (lihat bpr_pipeline.py)."""
    m = _RAW_TS_RE.search(source_name)
    if not m:
        return source_name
    ts = m.group(1)
    y, mo, d, hh, mm = int(ts[:4]), int(ts[4:6]), int(ts[6:8]), int(ts[8:10]), int(ts[10:12])
    return f"UPDATE {d} {_MONTH_ID[mo]} {y} {hh:02d}:{mm:02d}"


def _is_prev_total_col(col: str) -> bool:
    """'TOTAL 24 Ags 2026' (label dinamis dari previous_total_label()) -- beda
    dari kolom 'TOTAL' biasa, jadi dicek via prefix bukan exact match."""
    return col.startswith("TOTAL ") and col != "TOTAL"


def _col_fill(col: str) -> str:
    if col in ("NORM", "STOK", "DOI"):
        return C_DATA
    if col in BRAND_GROUP_1:
        return C_BRAND_1
    if col in BRAND_GROUP_2:
        return C_BRAND_2
    if col == "TOTAL" or _is_prev_total_col(col):
        return C_TOTAL
    if col in ("ACTUAL ORDER", "%"):
        return C_DARK
    return C_WHITE


def _col_font(col: str) -> str:
    return C_WHITE if col in ("ACTUAL ORDER", "%") else "000000"


def _fmt_num(v) -> str:
    """Gaya accounting template asli: 0 desimal, '-' untuk nol/kosong, kurung
    untuk negatif -- BUKAN 2/1 desimal yang biasa dipakai di halaman lain."""
    if pd.isna(v) or v == 0:
        return "-"
    s = f"{abs(v):,.0f}".replace(",", ".")
    return f"({s})" if v < 0 else s


def _fmt_pct(v) -> str:
    if pd.isna(v):
        return "-"
    return f"{v * 100:.0f}%"


def _fmt_cell(col: str, v):
    return _fmt_pct(v) if col == "%" else _fmt_num(v)


def _cols(df: pd.DataFrame) -> tuple[list, list]:
    id_cols = [c for c in ["Wilayah", "Depo"] if c in df.columns]
    value_cols = [c for c in df.columns if c not in id_cols]
    return id_cols, value_cols


def _is_total_row(row) -> bool:
    return str(row.get("Depo", row.get("Wilayah", ""))).strip().upper().endswith("TOTAL")


def _header_groups(value_cols: list) -> list[tuple[str, list, str]]:
    """[(label_grup, [kolom...], warna_grup)] -- 'DATA' menaungi NORM/STOK/DOI,
    'PROPOSED ORDER' menaungi 10 brand + TOTAL, sama persis merge H3:R3 /
    D3:F3 di template (satu warna utuh per grup, warna sub-kolom cuma tampil
    di baris kedua -- itu juga perilaku asli Excel utk merged cell)."""
    groups = []
    data_cols = [c for c in ("NORM", "STOK", "DOI") if c in value_cols]
    if data_cols:
        groups.append(("DATA", data_cols, C_DATA))
    order_cols = [c for c in bp.BRAND_COLUMNS if c in value_cols] + (["TOTAL"] if "TOTAL" in value_cols else [])
    if order_cols:
        groups.append(("PROPOSED ORDER", order_cols, C_BRAND_1))
    handled = set(data_cols) | set(order_cols)
    # Kolom lain (ACTUAL ORDER, %, dan TOTAL <tanggal lalu> yang labelnya
    # dinamis dari previous_total_label()) -- masing-masing grup sendiri
    # (span 2 baris, tanpa sub-header), warna diambil dari _col_fill() supaya
    # otomatis benar apa pun nama kolomnya.
    for c in value_cols:
        if c not in handled:
            groups.append((c, [c], _col_fill(c)))
    return groups


def build_html_table(df: pd.DataFrame, title: str, update_label: str) -> str:
    """Tabel HTML header 2-baris + warna + format PERSIS template Excel."""
    id_cols, value_cols = _cols(df)
    groups = _header_groups(value_cols)

    def th(text, bg, fg="000000", colspan=1, rowspan=1):
        span = f' colspan="{colspan}"' if colspan > 1 else ""
        span += f' rowspan="{rowspan}"' if rowspan > 1 else ""
        return (
            f'<th{span} style="background:#{bg};color:#{fg};padding:4px 8px;white-space:nowrap;'
            f'border:1px solid #999;text-align:center;font-family:{FONT_NAME};">{text}</th>'
        )

    row1 = "<tr>"
    for c in id_cols:
        row1 += th(c.upper(), "FFFFFF", rowspan=2)
    for label, cols, color in groups:
        row1 += th(label, color, _col_font(cols[0]), colspan=len(cols))
    row1 += "</tr>"

    row2 = "<tr>"
    for label, cols, _ in groups:
        if len(cols) == 1 and cols[0] == label:
            continue  # ACTUAL ORDER / % -- sudah rowspan=2 di row1, tidak ada sel row2
        for c in cols:
            row2 += th(c, _col_fill(c), _col_font(c))
    row2 += "</tr>"

    body_rows = []
    for _, row in df.iterrows():
        is_total = _is_total_row(row)
        row_bg = "#EEEEEE" if is_total else "white"
        cells = ""
        for c in id_cols:
            v = row[c] if pd.notna(row[c]) else ""
            cells += (
                f'<td style="background:{row_bg};color:#000;padding:4px 8px;text-align:left;'
                f'border:1px solid #ccc;font-weight:bold;white-space:nowrap;font-family:{FONT_NAME};">{v}</td>'
            )
        for c in value_cols:
            cells += (
                f'<td style="background:{row_bg};color:#000;padding:4px 8px;text-align:right;'
                f'border:1px solid #ccc;font-weight:bold;white-space:nowrap;font-family:{FONT_NAME};">'
                f'{_fmt_cell(c, row[c])}</td>'
            )
        body_rows.append(f"<tr>{cells}</tr>")

    return f"""
    <div style="overflow-x:auto; border:1px solid #999; border-radius:4px; font-family:{FONT_NAME};">
      <div style="padding:8px 10px 0 10px; font-weight:bold; font-size:1.8rem;">BPR BIA</div>
      <div style="padding:0 10px 8px 10px; font-weight:bold; font-size:1.1rem;">{update_label}</div>
      <div style="padding:0 10px 6px 10px; font-size:0.85rem; color:#888;">{title}</div>
      <table style="border-collapse:collapse; font-size:0.8rem; width:100%;">
        <thead>{row1}{row2}</thead>
        <tbody>{"".join(body_rows)}</tbody>
      </table>
    </div>
    """


def build_excel_bytes(depo_df: pd.DataFrame, wilayah_df: pd.DataFrame, update_label: str) -> bytes:
    """Excel dengan warna/format/header gabung PERSIS template asli -- 2 sheet,
    sama seperti 'Rekap Per DEPO' + 'Rekap Per WILAYAH' di file sumbernya."""
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    BOLD = Font(name="Calibri", bold=True, size=11)

    def write_sheet(ws, df: pd.DataFrame, sheet_title: str):
        ws.title = sheet_title
        ws.sheet_view.showGridLines = False
        id_cols, value_cols = _cols(df)
        groups = _header_groups(value_cols)

        ws.cell(row=1, column=1, value="BPR BIA").font = Font(name="Calibri", bold=True, size=28)
        ws.cell(row=2, column=1, value=update_label).font = Font(name="Calibri", bold=True, size=18)
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 23.5
        ws.row_dimensions[4].height = 43.5

        row1, row2 = 3, 4
        col_i = 1
        for c in id_cols:
            ws.merge_cells(start_row=row1, start_column=col_i, end_row=row2, end_column=col_i)
            cell = ws.cell(row=row1, column=col_i, value=c.upper())
            cell.font = BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for r in (row1, row2):
                ws.cell(row=r, column=col_i).border = border
            col_i += 1

        col_positions: dict[str, int] = {}
        for label, cols, color in groups:
            is_tall = len(cols) == 1 and cols[0] == label
            start_col = col_i
            for c in cols:
                col_positions[c] = col_i
                col_i += 1
            end_col = col_i - 1
            if is_tall:
                ws.merge_cells(start_row=row1, start_column=start_col, end_row=row2, end_column=end_col)
            else:
                ws.merge_cells(start_row=row1, start_column=start_col, end_row=row1, end_column=end_col)
            top_cell = ws.cell(row=row1, column=start_col, value=label)
            top_cell.font = Font(name="Calibri", bold=True, color=_col_font(cols[0]))
            top_cell.fill = PatternFill("solid", fgColor=color)
            top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for cc in range(start_col, end_col + 1):
                ws.cell(row=row1, column=cc).border = border
                ws.cell(row=row2, column=cc).border = border
            if not is_tall:
                for c in cols:
                    cell = ws.cell(row=row2, column=col_positions[c], value=c)
                    cell.font = Font(name="Calibri", bold=True, color=_col_font(c))
                    cell.fill = PatternFill("solid", fgColor=_col_fill(c))
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for r_off, (_, row) in enumerate(df.iterrows(), start=row2 + 1):
            is_total = _is_total_row(row)
            for i, c in enumerate(id_cols, start=1):
                cell = ws.cell(row=r_off, column=i, value=None if pd.isna(row[c]) else row[c])
                cell.font = BOLD
                cell.border = border
                if is_total:
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")
            for c, j in col_positions.items():
                v = row[c]
                cell = ws.cell(row=r_off, column=j, value=None if pd.isna(v) else float(v))
                cell.font = BOLD
                cell.border = border
                cell.number_format = "0%" if c == "%" else '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                if is_total:
                    cell.fill = PatternFill("solid", fgColor="EEEEEE")

        ws.column_dimensions[get_column_letter(1)].width = 25.4 if "Wilayah" in id_cols else 25.4
        if "Depo" in id_cols:
            ws.column_dimensions[get_column_letter(2)].width = 22.9
        for c, j in col_positions.items():
            ws.column_dimensions[get_column_letter(j)].width = 18 if c in bp.BRAND_COLUMNS else 11

    write_sheet(wb.active, depo_df, "Rekap Per DEPO")
    write_sheet(wb.create_sheet(), wilayah_df, "Rekap Per WILAYAH")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hex_rgb(h: str):
    h = h.lstrip("#")
    return tuple(int(h[i:i + 2], 16) / 255 for i in (0, 2, 4))


# Ukuran halaman PERSIS PDF resmi (BPR BIA <tgl> 2026.pdf, dicek langsung
# lewat PyMuPDF: page.rect = 792x612pt = 11x8.5in -- Letter landscape) --
# BUKAN halaman custom-lebar seperti percobaan sebelumnya. Ini kunci
# sebenarnya kenapa versi resminya kebaca normal di PDF viewer manapun:
# ukuran halamannya STANDAR, jadi viewer tidak perlu scale-down apa pun:
# font 6,5pt di kertas 11x8,5in itu memang ukuran normal buat tabel padat,
# bukan "kekecilan" -- masalah sebelumnya BUKAN font-nya, tapi halamannya
# yang custom-lebar (sampai 36in) sehingga viewer WAJIB scale-down banyak.
PAGE_W_IN, PAGE_H_IN = 11.0, 8.5
MARGIN_IN = 0.35
FONT_TITLE, FONT_SUBTITLE, FONT_SECTION = 16.5, 10.5, 9
FONT_HEADER, FONT_CELL = 6.5, 6.5

# Header multi-kata (mis. "PROST RAJAWALI LAGER") di PDF resmi DI-WRAP satu
# kata per baris (dicek langsung lewat koordinat span PyMuPDF -- "PROST ",
# "RAJAWALI ", "LAGER" masing-masing baris terpisah) -- itu yang bikin kolom
# brand bisa tetap sempit tanpa kepotong. Kolom yang cuma butuh selebar KATA
# TERPANJANG dalam label yang sudah di-wrap, bukan seluruh frasa.
def _wrap(label: str) -> str:
    return "\n".join(label.split(" "))


_PAD_IN = 0.09

if MATPLOTLIB_AVAILABLE:
    from matplotlib.font_manager import FontProperties
    from matplotlib.textpath import TextPath
    _BOLD_FP = FontProperties(weight="bold")


def _line_width_in(line: str, fontsize: float) -> float:
    """Lebar GLYPH SEBENARNYA (bukan tebakan per-karakter) -- ukuran karakter
    bold beda-beda lebarnya ("I" vs "W"), tebakan rata-rata sempat bikin
    WILAYAH/DEPO panjang (mis. "39-JATIM SELATAN HOREKA") tumpah ke kolom
    sebelah. TextPath render path glyph sesuai font/ukuran yang BENERAN
    dipakai di ax.text(), jadi lebarnya presisi."""
    if not line:
        return 0.0
    tp = TextPath((0, 0), line, size=fontsize, prop=_BOLD_FP)
    return tp.get_extents().width / 72.0


def _text_width_in(text: str, fontsize: float = FONT_CELL) -> float:
    """Lebar (inci) baris TERPANJANG dalam `text` (boleh multi-baris kalau
    sudah di-_wrap()) -- dasar utama lebar kolom dinamis di bawah."""
    longest = max((_line_width_in(line, fontsize) for line in text.split("\n")), default=0.0)
    return longest + _PAD_IN


def _col_width_in(df: pd.DataFrame, col: str, header_text: str) -> float:
    """Lebar kolom = selebar konten/header TERPANJANG yang benar-benar akan
    dicetak (bukan bobot tebakan) -- dijamin tidak pernah kepotong/tabrakan
    apa pun isi datanya (nama wilayah/depo beda-beda panjang antar baris)."""
    header_w = _text_width_in(header_text)
    if col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            data_w = max((_text_width_in(_fmt_cell(col, v)) for v in df[col]), default=0)
        else:
            data_w = max((_text_width_in(str(v)) for v in df[col].dropna()), default=0)
    else:
        data_w = 0
    return max(header_w, data_w)


def _build_table_figure(df: pd.DataFrame, title: str, update_label: str) -> "Figure":
    id_cols, value_cols = _cols(df)
    groups = _header_groups(value_cols)
    cols = id_cols + value_cols
    n_data_rows = len(df)
    header_h = 2.6  # baris sub-header lebih tinggi dari baris data -- muat teks 3-baris ter-wrap
    n_rows = n_data_rows + header_h

    fig = Figure(figsize=(PAGE_W_IN, PAGE_H_IN), dpi=150)
    FigureCanvasAgg(fig)
    fig.text(0.02, 0.965, "BPR BIA", fontsize=FONT_TITLE, fontweight="bold")
    fig.text(0.02, 0.925, update_label, fontsize=FONT_SUBTITLE, fontweight="bold")
    fig.text(0.02, 0.895, title, fontsize=FONT_SECTION, color="#666")

    # Area tabel: sisa halaman di bawah judul, dengan margin kiri/kanan/bawah kecil.
    table_top = 0.87
    ax = fig.add_axes([
        MARGIN_IN / PAGE_W_IN, MARGIN_IN / PAGE_H_IN,
        1 - 2 * MARGIN_IN / PAGE_W_IN, table_top - MARGIN_IN / PAGE_H_IN,
    ])

    header_label_of = {c: c.upper() for c in id_cols}
    for label, gcols, _ in _header_groups(value_cols):
        is_tall = len(gcols) == 1 and gcols[0] == label
        if is_tall:
            header_label_of[gcols[0]] = _wrap(label)
        else:
            for c in gcols:
                header_label_of[c] = _wrap(c)

    col_w_in = {c: _col_width_in(df, c, header_label_of[c]) for c in cols}
    total_w_in = sum(col_w_in.values())
    avail_w_in = PAGE_W_IN - 2 * MARGIN_IN
    # Kalau total kebutuhan pas/lebih dari lebar halaman, skala turun seragam
    # (jarang kejadian, tapi jangan biarkan kolom tumpah keluar halaman) --
    # kalau lebih kecil dari lebar halaman, biarkan longgar apa adanya (tidak
    # dipaksa nutupin sisa ruang, sama seperti tabel resminya yang tidak
    # memenuhi 100% lebar kertas).
    scale = min(1.0, avail_w_in / total_w_in) if total_w_in > 0 else 1.0
    # Kolom di-skala turun via `scale`, tapi ax.text() fontsize itu ukuran FISIK
    # tetap (pt), tidak ikut skala sumbu data -- kalau tidak font ikut dikecilkan
    # proporsional juga, teks (diukur lebar TextPath di font penuh) akan tumpah
    # keluar kolom yang sudah mengecil (lebih sempit dari lebar teks aslinya).
    font_header = FONT_HEADER * scale
    font_cell = FONT_CELL * scale

    ax.set_xlim(0, total_w_in * scale)
    ax.set_ylim(0, n_rows)
    ax.invert_yaxis()
    ax.axis("off")

    col_x = [0.0]
    for c in cols:
        col_x.append(col_x[-1] + col_w_in[c] * scale)

    rects, colors = [], []
    # baris 0: grup header (tinggi 1), baris 1: sub header (tinggi header_h-1,
    # muat teks ter-wrap) -- id_cols & tall groups span keduanya sekaligus.
    ci = 0
    for c in id_cols:
        rects.append(Rectangle((col_x[ci], 0), col_x[ci + 1] - col_x[ci], header_h))
        colors.append((1, 1, 1))
        ci += 1
    for label, gcols, color in groups:
        is_tall = len(gcols) == 1 and gcols[0] == label
        span = len(gcols)
        rgb = _hex_rgb(color)
        if is_tall:
            rects.append(Rectangle((col_x[ci], 0), col_x[ci + span] - col_x[ci], header_h))
            colors.append(rgb)
        else:
            rects.append(Rectangle((col_x[ci], 0), col_x[ci + span] - col_x[ci], 1))
            colors.append(rgb)
            for k, c in enumerate(gcols):
                rects.append(Rectangle((col_x[ci + k], 1), col_x[ci + k + 1] - col_x[ci + k], header_h - 1))
                colors.append(_hex_rgb(_col_fill(c)))
        ci += span

    for i, (_, row) in enumerate(df.iterrows()):
        ri = header_h + i
        bg = (0.93, 0.93, 0.93) if _is_total_row(row) else (1, 1, 1)
        for c in range(len(cols)):
            rects.append(Rectangle((col_x[c], ri), col_x[c + 1] - col_x[c], 1))
            colors.append(bg)

    ax.add_collection(PatchCollection(rects, facecolor=colors, edgecolor="#999", linewidth=0.3))

    ci = 0
    for c in id_cols:
        ax.text((col_x[ci] + col_x[ci + 1]) / 2, header_h / 2, header_label_of[c], ha="center", va="center",
                 fontsize=font_header, fontweight="bold")
        ci += 1
    for label, gcols, color in groups:
        is_tall = len(gcols) == 1 and gcols[0] == label
        span = len(gcols)
        fg = "white" if label in ("ACTUAL ORDER", "%") else "black"
        y = header_h / 2 if is_tall else 0.5
        ax.text((col_x[ci] + col_x[ci + span]) / 2, y, header_label_of[gcols[0]] if is_tall else label,
                 ha="center", va="center", fontsize=font_header, fontweight="bold", color=fg)
        if not is_tall:
            for k, c in enumerate(gcols):
                fg2 = "white" if c in ("ACTUAL ORDER", "%") else "black"
                ax.text((col_x[ci + k] + col_x[ci + k + 1]) / 2, 1 + (header_h - 1) / 2, header_label_of[c],
                         ha="center", va="center", fontsize=font_header, fontweight="bold", color=fg2)
        ci += span

    for i, (_, row) in enumerate(df.iterrows()):
        ri = header_h + i
        for c_i, c in enumerate(cols):
            if c in id_cols:
                text = "" if pd.isna(row[c]) else str(row[c])
                align = "left"
                x = col_x[c_i] + 0.05
            else:
                text = _fmt_cell(c, row[c])
                align = "right"
                x = col_x[c_i + 1] - 0.05
            ax.text(x, ri + 0.5, text, ha=align, va="center", fontsize=font_cell, fontweight="bold")

    return fig


def build_pdf_bytes(depo_df: pd.DataFrame, wilayah_df: pd.DataFrame, update_label: str) -> bytes:
    """PDF 2 halaman (Rekap Per Depo, Rekap Per Wilayah), ukuran & skala font
    PERSIS PDF resmi (Letter landscape, lihat _build_table_figure()) -- render
    via matplotlib (pola sama seperti render_outlet_image.py: Figure +
    PatchCollection, savefig ke format 'pdf' alih-alih 'png')."""
    if not MATPLOTLIB_AVAILABLE:
        raise RuntimeError(f"matplotlib tidak bisa dimuat: {_MATPLOTLIB_IMPORT_ERROR}")

    from matplotlib.backends.backend_pdf import PdfPages

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        for df, title in [(depo_df, "Rekap Per Depo"), (wilayah_df, "Rekap Per Wilayah")]:
            fig = _build_table_figure(df, title, update_label)
            pdf.savefig(fig)
    return buf.getvalue()
