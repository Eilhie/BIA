r"""
CUKAI PIPELINE
Baca data volume kompetitor (estimasi dari data cukai/pajak minuman beralkohol
-- satu-satunya cara mengetahui volume kompetitor karena angka penjualan
mereka sendiri tidak pernah kita punya) dari file kerja analis di
D:\cukai kompetitor. TIDAK menulis apa pun ke file itu -- baca & tampilkan
saja, sama seperti pola Gabungan HOREKA.

Cover sheet 'REKAP RAPIH' (tabel bulanan bersih "Bir OT vs Musuh" per brand)
+ sheet detail per kompetitor (Balihai/Delta, Lokal/Export -- lihat
DETAIL_SHEETS) yang pecah sampai level kemasan/SKU. Folder ini juga punya
file level-wilayah (FORMAT CUKAI WILAYAH.xlsx) dan data mentah per-brand
yang BELUM di-cover -- strukturnya jauh lebih kompleks (header 4 tingkat,
kemungkinan multi-blok), disengaja skip dulu.
"""

import datetime
import re
from pathlib import Path

import openpyxl
import pandas as pd

CUKAI_DIR = Path(r"D:\cukai kompetitor")
REKAP_FILE = CUKAI_DIR / "Cukai Kompetitor.xlsx"
REKAP_SHEET = "REKAP RAPIH"

MONTH_ID = ["JAN", "FEB", "MAR", "APR", "MEI", "JUN", "JUL", "AGS", "SEP", "OKT", "NOV", "DES"]


def is_available() -> bool:
    return REKAP_FILE.exists()


def load_rekap_rapih() -> pd.DataFrame:
    """Parse REKAP RAPIH jadi long-format (Brand, Month, Qty, Share) -- kolom bulan
    dideteksi otomatis lewat header row 2 yang selnya bertipe datetime (bukan
    hardcode posisi kolom), supaya tahan kalau baris/kolom quarter/rata2 di
    sekitarnya berubah. Tiap kolom bulan itu QTY-nya persis di kolom itu, share
    (pangsa, 0.0-1.0) di kolom SEBELAHNYA -- pola ini konsisten dicek lewat
    merged-cell range file aslinya (mis. D2:E2 utk Jan25 = kolom D:qty, E:share)."""
    if not is_available():
        return pd.DataFrame(columns=["Brand", "Month", "MonthKey", "Qty", "Share"])

    wb = openpyxl.load_workbook(REKAP_FILE, data_only=True, read_only=True)
    ws = wb[REKAP_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return pd.DataFrame(columns=["Brand", "Month", "MonthKey", "Qty", "Share"])

    header = rows[1]
    month_cols = [i for i, v in enumerate(header) if isinstance(v, datetime.datetime)]

    records = []
    for row in rows[2:]:
        brand = row[0]
        if not brand or str(brand).strip() == "" or str(brand).strip().lower() == "total":
            continue
        brand = str(brand).strip()
        for c in month_cols:
            month_dt = header[c]
            qty = row[c] if c < len(row) else None
            share = row[c + 1] if c + 1 < len(row) else None
            if qty is None and share is None:
                continue
            month_key = f"{MONTH_ID[month_dt.month - 1]}{month_dt.year % 100:02d}"
            records.append({
                "Brand": brand,
                "Month": month_dt.strftime("%b %y"),
                "MonthKey": month_key,
                "MonthDate": month_dt,
                "Qty": qty or 0,
                "Share": share,
            })

    return pd.DataFrame(records)


def get_last_modified() -> datetime.datetime | None:
    if not is_available():
        return None
    return datetime.datetime.fromtimestamp(REKAP_FILE.stat().st_mtime)


# ── SHEET DETAIL LOKAL/EXPORT (Balihai, Delta) ────────────────────────────────

_MONTH_ABBR_EN = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
_MONTH_LABEL_RE = re.compile(r"^([A-Za-z]{3})[a-z]*\s+(\d{2,4})$")

DETAIL_SHEETS = {
    "Bali Hai - Lokal": "Balihai Lokal",
    "Bali Hai - Export": "Balihai Export",
    "Delta - Lokal": "Delta Lokal",
    "Delta - Export": "Delta Export",
}


def _parse_month_header_cell(v) -> tuple[int, int] | None:
    """'Jan 24' / datetime(2024,1,1) -> (2024, 1); None kalau bukan label bulan
    (mis. 'Q1 2024', header lain) -- dipakai buat saring KOLOM QUARTER/tahunan
    yang ikut lolos filter 'QTY KRT' di baris sub-header tapi bukan bulan asli."""
    if isinstance(v, datetime.datetime):
        return v.year, v.month
    if isinstance(v, str):
        m = _MONTH_LABEL_RE.match(v.strip())
        if m:
            mon = _MONTH_ABBR_EN.get(m.group(1).lower())
            if mon:
                yy = int(m.group(2))
                year = yy + 2000 if yy < 100 else yy
                return year, mon
    return None


def load_detail_sheet(sheet_key: str) -> pd.DataFrame:
    """Parse sheet detail Lokal/Export (Balihai, Delta) -- struktur beda dari
    REKAP RAPIH: baris label ada di kolom B (bukan A), sub-header 'QTY KRT'/
    'QTY LITER' berpasangan per kolom bulan, dan baris berjenjang: baris TOTAL
    per brand/varian (font BOLD) diikuti baris detail kemasan/SKU (font biasa)
    di bawahnya -- font.bold jadi penanda paling andal (dicek langsung ke file,
    kolom A kadang keisi kadang tidak, tidak konsisten antar sheet)."""
    sheet_name = DETAIL_SHEETS.get(sheet_key)
    if sheet_name is None or not is_available():
        return pd.DataFrame(columns=["Baris", "IsTotal", "Month", "MonthKey", "MonthDate", "Qty", "Liter"])

    wb = openpyxl.load_workbook(REKAP_FILE, data_only=True)
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame(columns=["Baris", "IsTotal", "Month", "MonthKey", "MonthDate", "Qty", "Liter"])
    ws = wb[sheet_name]

    # cari baris sub-header 'QTY KRT' (selalu 1 baris di bawah baris label bulan)
    header_row = None
    for r in range(1, 10):
        if any(str(ws.cell(row=r, column=c).value).strip() == "QTY KRT" for c in range(1, ws.max_column + 1)):
            header_row = r
            break
    if header_row is None:
        return pd.DataFrame(columns=["Baris", "IsTotal", "Month", "MonthKey", "MonthDate", "Qty", "Liter"])

    month_col_label_row = header_row - 1
    month_cols = []  # (col_idx, year, month)
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=header_row, column=c).value).strip() != "QTY KRT":
            continue
        parsed = _parse_month_header_cell(ws.cell(row=month_col_label_row, column=c).value)
        if parsed:
            month_cols.append((c, *parsed))

    def _num(v):
        # sebagian sel di sumbernya berisi error formula literal (mis. '#N/A' dari
        # VLOOKUP yang gagal, ketemu nyata di sheet Bali Hai Export mulai Apr 26) --
        # bukan angka, jangan dipaksa 0 (menyembunyikan masalah), biarkan NaN supaya
        # kelihatan sebagai "-" di tabel, beda dari 0 yang genuinely nol.
        if isinstance(v, (int, float)):
            return v
        return None

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        label_cell = ws.cell(row=r, column=2)
        label = label_cell.value
        if not label or not str(label).strip():
            continue
        is_total = bool(label_cell.font.bold)
        for c, year, month in month_cols:
            qty = ws.cell(row=r, column=c).value
            liter = ws.cell(row=r, column=c + 1).value
            if qty is None and liter is None:
                continue
            records.append({
                "Baris": str(label).strip(),
                "IsTotal": is_total,
                "Month": f"{[k for k,v in _MONTH_ABBR_EN.items() if v==month][0].capitalize()} {year % 100:02d}",
                "MonthKey": f"{MONTH_ID[month - 1]}{year % 100:02d}",
                "MonthDate": datetime.datetime(year, month, 1),
                "Qty": _num(qty),
                "Liter": _num(liter),
            })

    return pd.DataFrame(records)
