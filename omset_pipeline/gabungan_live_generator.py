r"""
GABUNGAN LIVE GENERATOR
Kalau file Gabungan UMUM terbaru yang tersedia adalah format lama ("Toko
Gabungan Update <tgl> <bulan> <tahun>.xlsx") dan BELUM ada padanan "Formula
Live"-nya untuk bulan yang sama, script ini membuat satu -- struktur outlet/
grup disalin apa adanya dari file sumber, ditambah formula VLOOKUP live per
baris toko yang menunjuk ke D:\DB OMSHAR\DB\OMSHAR UMUM BIR.xls (file yang
SUDAH kita sync sendiri), meniru persis mekanisme file "Formula Live" asli
dari divisi lain (dibongkar & diverifikasi manual sebelum script ini ditulis
-- lihat externalLink1.xml / VLOOKUP pattern di percakapan).

PENTING -- output TIDAK ditulis ke D:\Data BIA\INFO BIA\Toko Gabungan\ (folder
sumber milik divisi lain) -- ditulis ke omset_pipeline/output/GABUNGAN_LIVE/
di tree kita sendiri, supaya tidak pernah tercampur/disangka file asli mereka.
Kalau mau dipakai beneran, pindahkan manual ke folder sumber setelah dicek.

Element order di xl/workbook.xml itu STRICT (schema CT_Workbook): sheets ->
externalReferences -> definedNames -> calcPr. Excel MENOLAK file kalau ini
salah urutan (openpyxl sendiri lebih toleran saat baca, makanya bug ini
tidak ketahuan sampai dicoba buka pakai Excel COM asli -- lihat riwayat
percakapan, tes pertama gagal persis karena ini)."""

import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import omset_seeker as os_

OUTPUT_DIR = Path(__file__).resolve().parent / "output" / "GABUNGAN_LIVE"
BIR_XLS_ABS = Path(r"D:\DB OMSHAR\DB\OMSHAR UMUM BIR.xls")
# Target relationship path -- URL-encoded absolute path style, persis format
# yang dipakai file "Formula Live" asli (lihat externalLink1.xml.rels-nya).
BIR_EXTERNAL_TARGET = "/DB%20OMSHAR/DB/OMSHAR%20UMUM%20BIR.xls"

WILAYAH_SHEETS = ["BTN", "DKI", "BDB", "JBS", "JBU", "JTU", "JTS", "JIU", "JIS", "BLI",
                  "SMU", "SMB", "SMS", "LPB", "NTR", "KLT", "KLB", "SLS", "SLU", "PPA"]

# Kolom VLOOKUP di file asli: G..U (15 kolom), col_index 140..154 di sheet
# wilayah BIR.xls -- disalin persis dari formula nyata yang ditemukan, BUKAN
# ditebak (lihat percakapan: G3..U3 = VLOOKUP(..., 140..154, FALSE)).
VLOOKUP_COLS = list("GHIJKLMNOPQRSTU")  # 15 kolom
VLOOKUP_COL_INDEX_START = 140


def _needs_generation() -> tuple[Path, int, int] | None:
    """Return (source_path, year, month) kalau file terbaru yang ada adalah
    format 'Update' (bukan 'Formula Live') DAN belum ada Formula Live untuk
    bulan yang sama -- None kalau tidak perlu generate apa-apa."""
    path = os_.find_latest_toko_gabungan()
    if path is None:
        return None
    m = os_._TOKO_GABUNGAN_NAME_RE.match(path.name)
    if not m:
        return None  # yang terbaru sudah Formula Live, atau nama tidak dikenali -- tidak perlu apa-apa
    day, month_name, year = m.groups()
    month = os_._INDO_MONTHS.get(month_name.lower())
    if month is None:
        return None

    live_re = os_._TOKO_GABUNGAN_LIVE_NAME_RE
    for p in os_.TOKO_GABUNGAN_DIR.glob("Toko Gabungan (Formula Live)*.xlsx"):
        lm = live_re.match(p.name)
        if not lm:
            continue
        lmonth = os_._INDO_MONTHS.get(lm.group(1).lower())
        if lmonth == month and int(lm.group(2)) == year:
            return None  # sudah ada Formula Live untuk bulan ini
    return path, year, month


def _read_raw_rows(path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _block_wilayah_per_row(rows: list[tuple]) -> list[str | None]:
    """Untuk tiap baris, cari wilayah blok yang menaunginya -- diambil dari baris
    RINGKASAN yang menutup blok itu (kolom A terisi, lihat _gabungan_from_block_sheet
    di omset_seeker.py buat pola yang sama). Baris kosong/ringkasan sendiri -> None."""
    result: list[str | None] = [None] * len(rows)
    block_start = None
    for i, row in enumerate(rows):
        is_blank = all(v is None for v in row)
        if is_blank:
            block_start = None
            continue
        if block_start is None:
            block_start = i
        if row[0] is not None:  # baris ringkasan -- menutup blok
            wilayah = row[0]
            for j in range(block_start, i):  # anak-anaknya, BUKAN baris ringkasan sendiri
                result[j] = wilayah
            block_start = None
    return result


def generate(source_path: Path, year: int, month: int) -> Path:
    """Bangun file Formula Live baru dari source_path, simpan ke OUTPUT_DIR.
    Return path file yang dihasilkan."""
    rows = _read_raw_rows(source_path)
    wilayah_per_row = _block_wilayah_per_row(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "A11"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=r_idx, column=c_idx, value=val)

        site = row[2] if len(row) > 2 else None
        is_summary = row[0] is not None
        wilayah = wilayah_per_row[r_idx - 1]
        if site and not is_summary and wilayah in WILAYAH_SHEETS:
            for col_letter, col_idx in zip(VLOOKUP_COLS, range(VLOOKUP_COL_INDEX_START, VLOOKUP_COL_INDEX_START + len(VLOOKUP_COLS))):
                formula = f'=VLOOKUP("*"&RIGHT($C{r_idx},8),[1]{wilayah}!$B$1:$FT$65536,{col_idx},FALSE)'
                ws[f"{col_letter}{r_idx}"] = formula

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    month_name = [k for k, v in os_._INDO_MONTHS.items() if v == month][0].capitalize()
    out_path = OUTPUT_DIR / f"Toko Gabungan (Formula Live) - Data {month_name} {year}.xlsx"
    tmp_path = out_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)

    _inject_external_link(tmp_path)

    tmp_path.replace(out_path)
    return out_path


def _inject_external_link(path: Path) -> None:
    """Suntik xl/externalLinks/externalLink1.xml + relationship + registrasi di
    workbook.xml/rels/[Content_Types].xml -- diverifikasi manual pakai Excel COM
    beneran (bukan cuma openpyxl, yang lebih toleran) sebelum dipakai di sini."""
    sheet_names_xml = "".join(f'<sheetName val="{w}"/>' for w in WILAYAH_SHEETS)
    external_link1_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<externalBook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1">'
        f'<sheetNames>{sheet_names_xml}</sheetNames></externalBook></externalLink>'
    )
    external_link1_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" '
        f'Target="{BIR_EXTERNAL_TARGET}" TargetMode="External"/></Relationships>'
    )

    tmp = path.with_name(f".{path.name}.zipwork.tmp")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/workbook.xml":
                text = data.decode("utf-8")
                assert "</sheets>" in text, "workbook.xml tidak punya </sheets> -- struktur openpyxl berubah?"
                text = text.replace(
                    "</sheets>",
                    '</sheets><externalReferences><externalReference r:id="rIdExt1"/></externalReferences>',
                )
                data = text.encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                text = text.replace(
                    "</Relationships>",
                    '<Relationship Id="rIdExt1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink" '
                    'Target="externalLinks/externalLink1.xml"/></Relationships>',
                )
                data = text.encode("utf-8")
            elif item.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                text = text.replace(
                    "</Types>",
                    '<Override PartName="/xl/externalLinks/externalLink1.xml" '
                    'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"/></Types>',
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
        zout.writestr("xl/externalLinks/externalLink1.xml", external_link1_xml)
        zout.writestr("xl/externalLinks/_rels/externalLink1.xml.rels", external_link1_rels)
    tmp.replace(path)


if __name__ == "__main__":
    need = _needs_generation()
    if need is None:
        print("Tidak perlu generate -- Formula Live untuk bulan ini sudah ada, atau file terbaru sudah Formula Live.")
    else:
        source_path, year, month = need
        print(f"Generating dari: {source_path.name}")
        out = generate(source_path, year, month)
        print(f"Selesai: {out}")
