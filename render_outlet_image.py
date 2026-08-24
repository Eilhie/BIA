"""
Render laporan OMSET per outlet jadi gambar (PNG), meniru layout OMSET OUTLET asli:
  - Header: kode site (boxed), nama outlet, cut off date
  - Tabel: 12 bulan 2025 + RT2 25 | BRAND | 12 bulan 2026 + RT2 26
  - Baris brand (19 SKU) + BIR (total kategori) + BEV (DIV AB1 - BIR) + DIV AB1
  - Footer: hierarki wilayah (Propinsi/Kota/Kecamatan/Alamat)
"""

import io
import os
import time
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

try:
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.collections import PatchCollection
    from matplotlib.figure import Figure
    from matplotlib.patches import Rectangle
    MATPLOTLIB_AVAILABLE = True
    _MATPLOTLIB_IMPORT_ERROR = None
except ImportError as _e:
    # mis. matplotlib diblokir Windows Smart App Control -- jangan biarkan itu
    # menjatuhkan seluruh modul (dan modul lain yang import build_report_rows
    # dari sini, yang sama sekali tidak butuh matplotlib). render_outlet_report()
    # akan raise error yang jelas saat benar-benar dipanggil, bukan saat import.
    MATPLOTLIB_AVAILABLE = False
    _MATPLOTLIB_IMPORT_ERROR = _e

from omset_seeker import (
    BRAND_ORDER,
    DIV_AB1_BRAND,
    HOREKA_KEG_BRAND_ORDER,
    VALIDATION_BRAND,
    compute_bev,
    get_brand_months,
    get_cutoff_date,
    get_cutoff_parts,
    seek_outlet,
)

OUTPUT_DIR = Path(__file__).resolve().parent / "omset_pipeline" / "output" / "IMAGE"

MONTHS_ID = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Ags", "Sep", "Okt", "Nov", "Des"]
LABELS_25 = [f"{m}-25" for m in MONTHS_ID]
LABELS_26 = [f"{m}-26" for m in MONTHS_ID]

COL_LABELS = LABELS_25 + ["RT2 25", "BRAND"] + LABELS_26 + ["RT2 26"]
COL_WIDTHS = [1.1] * 12 + [1.1, 2.6] + [1.1] * 12 + [1.1]
N_COLS = len(COL_LABELS)
BRAND_COL_IDX = 13  # index posisi kolom "BRAND" di COL_LABELS

# Warna
C_HEADER_25 = "#FCE0C3"
C_HEADER_RT2_25 = "#E8281A"
C_HEADER_BRAND = "#F8E2A8"
C_HEADER_26 = "#FFFFFF"
C_HEADER_RT2_26 = "#1F6FB2"
C_CELL_RT2_25 = "#FBE3E1"
C_CELL_RT2_26 = "#DCEAF6"
C_CELL_BRAND = "#EAF1FB"
C_TOTAL_ROW = "#FCD9A0"   # BIR
C_BEV_ROW = "#DDEBF7"     # BEV
C_DIVAB1_ROW = "#3D5A80"  # DIV AB1 -- digelapkan supaya teks putih di atasnya kontras jelas
C_GRID = "#BFBFBF"

ROW_BG = {"normal": "white", "bir": C_TOTAL_ROW, "bev": C_BEV_ROW, "divab1": C_DIVAB1_ROW}


def fmt_krt(v) -> str:
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "-"
    if v == 0:
        return "-"
    s = f"{v:,.1f}"
    return s.replace(",", "§").replace(".", ",").replace("§", ".")


def build_row_cells(month_values: dict, months_26_closed: int = 12) -> list:
    """months_26_closed = jumlah bulan 2026 yang sudah genap tutup buku (sebelum bulan cut off) --
    RT2 26 cuma dirata-rata dari bulan itu, bukan selalu /12, supaya tidak kekecilan
    gara-gara ikut membagi dengan bulan yang belum ada datanya sama sekali."""
    vals_25 = [fmt_krt(month_values.get(f"{m} 2025", 0)) for m in
               ["JAN", "FEB", "MAR", "APR", "MEI", "JUN", "JUL", "AGS", "SEP", "OKT", "NOV", "DES"]]
    vals_26 = [fmt_krt(month_values.get(f"{m} 2026", 0)) for m in
               ["JAN", "FEB", "MAR", "APR", "MEI", "JUN", "JUL", "AGS", "SEP", "OKT", "NOV", "DES"]]
    rt2_25 = sum(month_values.get(f"{m} 2025", 0) for m in
                 ["JAN", "FEB", "MAR", "APR", "MEI", "JUN", "JUL", "AGS", "SEP", "OKT", "NOV", "DES"]) / 12
    months_26_labels = ["JAN", "FEB", "MAR", "APR", "MEI", "JUN", "JUL", "AGS", "SEP", "OKT", "NOV", "DES"][:months_26_closed]
    if months_26_labels:
        rt2_26 = sum(month_values.get(f"{m} 2026", 0) for m in months_26_labels) / len(months_26_labels)
    else:
        rt2_26 = 0
    return vals_25, fmt_krt(rt2_25), vals_26, fmt_krt(rt2_26)


def build_report_rows(site: str, omshar_type: str = "UMUM", with_keg: bool = False):
    """Siapkan row_cells (vals_25, rt2_25, label, vals_26, rt2_26, row_type) + info + cutoff,
    tanpa render gambar -- dipakai render_outlet_report() (PNG) dan omset_search_app.py (tabel
    web) supaya format & logika RT2 26 selalu konsisten di kedua tempat."""
    brands = BRAND_ORDER + HOREKA_KEG_BRAND_ORDER if with_keg else BRAND_ORDER
    table, info = seek_outlet(site, omshar_type, brands=brands)
    if info is None:
        raise ValueError(f"Site '{site}' tidak ditemukan di {omshar_type}.")

    cutoff = get_cutoff_date(omshar_type=omshar_type)
    cutoff_parts = get_cutoff_parts(omshar_type=omshar_type)
    # bulan 2026 yang sudah genap tutup buku = sebelum bulan cut off (bulan cut off sendiri
    # masih berjalan, belum genap sebulan penuh, jadi dikecualikan dari RT2 26)
    months_26_closed = (cutoff_parts[1] - 1) if cutoff_parts else 12

    def make_row(label, month_values, row_type="normal"):
        vals_25, rt2_25, vals_26, rt2_26 = build_row_cells(month_values, months_26_closed)
        return (vals_25, rt2_25, label, vals_26, rt2_26, row_type)

    # Baris data: SKU (+ KEG jika with_keg) + BIR (total kategori) + BEV (DIV AB1 - BIR) + DIV AB1
    row_cells = [make_row(brand, table.loc[brand].to_dict()) for brand in brands]
    row_cells.append(make_row(VALIDATION_BRAND, get_brand_months(VALIDATION_BRAND, site, omshar_type), "bir"))
    row_cells.append(make_row("BEV", compute_bev(site, omshar_type), "bev"))
    row_cells.append(make_row(DIV_AB1_BRAND, get_brand_months(DIV_AB1_BRAND, site, omshar_type), "divab1"))

    return row_cells, info, cutoff


def build_html_table(row_cells: list, cutoffs: dict[str, str] | None = None) -> str:
    """Tabel HTML yang meniru layout & warna OMSET OUTLET Excel/PNG asli:
    2025 (peach) + RT2 25 (merah) | BRAND (biru muda) | CUT OFF (opsional) |
    2026 (putih) + RT2 26 (biru tua). Dipindah ke sini (dari omset_search_app.py)
    supaya bisa dipakai ulang di halaman lain (mis. Detail SKU Brand Besar) tanpa
    import omset_search_app.py -- itu skrip yang dieksekusi sebagai halaman
    Streamlit sendiri (ada st.title/sidebar dsb di top-level), TIDAK aman
    di-import sebagai modul biasa dari halaman lain.

    Kolom CUT OFF cuma tampil kalau `cutoffs` diisi -- SENGAJA tidak disentuh di
    render_outlet_report() (PNG) sama sekali, jadi PNG yang dipakai Print/Copy/
    Excel otomatis TIDAK ikut nampilkan kolom ini tanpa perlu CSS/logic khusus
    "sembunyikan saat print", karena keduanya memang jalur render yang benar-benar
    terpisah.

    Kolom BRAND di-freeze (position:sticky) ke tepi kiri area scroll -- tabelnya
    lebar (26 kolom), dan BRAND ada di posisi ke-14 (setelah 2025+RT2 25), jadi
    tanpa ini nama brand-nya ilang duluan begitu discroll ke kanan, terutama di
    tempat sempit kayak modal (lihat pages/5_Outlet_Lapisan_MClub.py)."""
    show_cutoff = cutoffs is not None
    STICKY = "position:sticky;left:0;z-index:2;box-shadow:2px 0 3px -1px rgba(0,0,0,0.3);"
    th = lambda text, bg, fg="black", extra="": (
        f'<th style="background:{bg};color:{fg};padding:4px 8px;white-space:nowrap;'
        f'border:1px solid #999;{extra}">{text}</th>'
    )
    header = "<tr>"
    header += "".join(th(c, C_HEADER_25) for c in LABELS_25)
    header += th("RT2 25", C_HEADER_RT2_25, "white")
    header += th("BRAND", C_HEADER_BRAND, extra=STICKY)
    if show_cutoff:
        header += th("CUT OFF", C_HEADER_BRAND)
    header += "".join(th(c, C_HEADER_26) for c in LABELS_26)
    header += th("RT2 26", C_HEADER_RT2_26, "white")
    header += "</tr>"

    body_rows = []
    for vals_25, rt2_25, label, vals_26, rt2_26, row_type in row_cells:
        bg = ROW_BG[row_type]
        fg = "white" if row_type == "divab1" else "black"
        td = lambda text, cell_bg, align="right", bold=False, extra="": (
            f'<td style="background:{cell_bg};color:{fg};padding:4px 8px;text-align:{align};'
            f'border:1px solid #ccc;{"font-weight:bold;" if bold else ""}white-space:nowrap;{extra}">{text}</td>'
        )
        row = "<tr>"
        row += "".join(td(v, bg) for v in vals_25)
        row += td(rt2_25, bg if row_type != "normal" else C_CELL_RT2_25, bold=True)
        row += td(label, bg if row_type != "normal" else C_CELL_BRAND, align="left", bold=True, extra=STICKY)
        if show_cutoff:
            row += td(cutoffs.get(label, "") or "-", bg, align="center")
        row += "".join(td(v, bg) for v in vals_26)
        row += td(rt2_26, bg if row_type != "normal" else C_CELL_RT2_26, bold=True)
        row += "</tr>"
        body_rows.append(row)

    return f"""
    <div style="overflow-x:auto; border:1px solid #999; border-radius:4px;">
      <table style="border-collapse:collapse; font-size:0.8rem; width:100%;">
        <thead>{header}</thead>
        <tbody>{"".join(body_rows)}</tbody>
      </table>
    </div>
    """


def _parse_krt(s: str) -> float:
    """Kebalikan dari fmt_krt() -- "1.234,5" -> 1234.5, "-" -> 0.0."""
    if s == "-":
        return 0.0
    return float(s.replace(".", "").replace(",", "."))


def build_report_excel(
    site: str | None = None,
    omshar_type: str = "UMUM",
    with_keg: bool = False,
    *,
    precomputed: tuple | None = None,
) -> bytes:
    """Bangun file .xlsx (bytes, siap di-download) dari data yang sama persis dengan
    tabel web/PNG. Reuse build_report_rows() lalu parse balik string terformat jadi
    angka asli (bukan hitung ulang dari nol) -- supaya RT2 26/BIR/BEV tidak pernah bisa
    beda dari yang ditampilkan di web atau PNG.

    `precomputed` = tuple (row_cells, info, cutoff) dari build_report_rows() yang sudah
    dihitung pemanggil -- dipakai omset_search_app.py supaya query data tidak diulang 3x
    per tampilan (tabel + Excel + PNG). Kalau None, hitung sendiri (jalur CLI)."""
    if precomputed is None:
        row_cells, info, cutoff = build_report_rows(site, omshar_type, with_keg)
    else:
        row_cells, info, cutoff = precomputed

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OMSET"

    ws.append([info["Site"]])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws.append([info["Outlet"]])
    ws.cell(row=2, column=1).font = Font(bold=True, size=12)

    ws.append([f"CUT OFF : {cutoff}"])
    ws.cell(row=3, column=1).font = Font(bold=True, color="D40000")

    ws.append([f"{info['Propinsi']} / {info['Kota']} / {info['Kecamatan']} / {info['Alamat']}"])
    ws.cell(row=4, column=1).font = Font(italic=True, color="666666")

    ws.append([])

    # C_GRID -- di PNG semua sel dikasih border tipis, jadi kolom putih (2026) tetap
    # kelihatan bentuk selnya. Excel tidak otomatis kasih border, jadi kalau tidak
    # ditambahkan manual, sel dengan fill putih nyatu sama background sheet dan
    # kelihatan kosong/tidak terformat.
    thin_gray = Side(style="thin", color=C_GRID.lstrip("#"))
    cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    header = LABELS_25 + ["RT2 25", "BRAND"] + LABELS_26 + ["RT2 26"]
    ws.append(header)
    header_row = ws.max_row
    for c, label in enumerate(header, start=1):
        if label in LABELS_25:
            fill = C_HEADER_25
        elif label == "RT2 25":
            fill = C_HEADER_RT2_25
        elif label == "BRAND":
            fill = C_HEADER_BRAND
        elif label == "RT2 26":
            fill = C_HEADER_RT2_26
        else:
            fill = C_HEADER_26
        cell = ws.cell(row=header_row, column=c)
        cell.fill = PatternFill(start_color=fill.lstrip("#"), end_color=fill.lstrip("#"), fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF" if label in ("RT2 25", "RT2 26") else "000000")
        cell.border = cell_border

    row_bg_hex = {k: (v.lstrip("#") if v != "white" else "FFFFFF") for k, v in ROW_BG.items()}
    for vals_25, rt2_25, brand, vals_26, rt2_26, row_type in row_cells:
        values = (
            [_parse_krt(v) for v in vals_25] + [_parse_krt(rt2_25)] + [brand]
            + [_parse_krt(v) for v in vals_26] + [_parse_krt(rt2_26)]
        )
        ws.append(values)
        r = ws.max_row
        fill_hex = row_bg_hex[row_type]
        font_color = "FFFFFF" if row_type == "divab1" else "000000"
        for c, (label, val) in enumerate(zip(header, values), start=1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            cell.border = cell_border
            is_brand_col = label == "BRAND"
            cell.font = Font(bold=is_brand_col, color=font_color)
            if not is_brand_col:
                cell.number_format = "#,##0.0"

    ws.column_dimensions["N"].width = 22  # kolom BRAND
    for c, label in enumerate(header, start=1):
        if label != "BRAND":
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 9

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_figure(row_cells: list, info: dict, cutoff: str) -> Figure:
    """Bangun figure matplotlib laporan OMSET (murni gambar, tanpa I/O).

    Optimalisasi: semua Rectangle sel (~650) digabung dalam SATU PatchCollection --
    path & urutan row-major sama, jumlah artist turun drastis sehingga build + savefig
    jauh lebih cepat (hasil raster praktis identik, lihat catatan di bagian grid)."""
    n_data_rows = len(row_cells)
    n_grid_rows = 1 + n_data_rows  # header + data

    # ── Layout koordinat ──
    x_edges = [0]
    for w in COL_WIDTHS:
        x_edges.append(x_edges[-1] + w)
    total_w = x_edges[-1]

    row_h = 1.0
    grid_top_y = 0
    grid_bottom_y = n_grid_rows * row_h

    header_text_h = row_h * 1.1
    footer_text_h = row_h * 0.9
    n_header_lines = 3
    n_footer_lines = 4

    fig_w = total_w * 0.5
    fig_h = (n_header_lines * header_text_h + n_grid_rows * row_h + n_footer_lines * footer_text_h) * 0.32
    # Figure() OO API, BUKAN plt.figure() -- pyplot menyimpan state figure aktif
    # secara GLOBAL (Gcf registry) yang tidak thread-safe. Streamlit menjalankan tiap
    # sesi browser di thread terpisah dalam proses yang sama; kalau 2 render jalan
    # bersamaan (mis. 2 tab), plt.figure()/plt.close() bisa saling nimpa/nyampur
    # antar figure di tengah proses gambar -- itu penyebab PNG "burik"/dobel
    # sebelumnya, BUKAN race saat nulis file (itu sudah dibenerin lewat atomic write
    # di bawah, tapi race race sebenarnya terjadi lebih awal, saat menggambar).
    fig = Figure(figsize=(fig_w, fig_h))
    FigureCanvasAgg(fig)
    ax = fig.add_axes([0, 0, 1, 1])

    top_y = -(n_header_lines * header_text_h)
    ax.set_xlim(0, total_w)
    ax.set_ylim(grid_bottom_y + n_footer_lines * footer_text_h, top_y)
    ax.axis("off")

    # ── Header text (site code, outlet, cutoff) ──
    cx = total_w / 2
    y = top_y + header_text_h * 0.5
    ax.text(cx, y, info["Site"], ha="center", va="center", fontsize=13, fontweight="bold")

    y += header_text_h
    ax.text(cx, y, info["Outlet"], ha="center", va="center", fontsize=14, fontweight="bold")

    y += header_text_h
    ax.text(cx, y, f"CUT OFF : {cutoff}", ha="center", va="center", fontsize=11,
            fontweight="bold", color="#D40000")

    # ── Latar + grid sel: tetap satu Rectangle per sel (bentuk & posisi sama persis
    # dengan add_patch satu-satu), tapi digabung dalam SATU PatchCollection -- jumlah
    # artist turun dari ~650 patch jadi 1, build + rasterisasi jauh lebih cepat.
    # Urutan path tetap row-major seperti semula supaya kompositing edge hampir
    # identik (selisih raster terukur cuma ~100 px anti-aliasing delta <=6 di sudut
    # garis tepi tabel, tidak terlihat mata). ──
    def header_bg(col_idx):
        if col_idx < 12:
            return C_HEADER_25
        if col_idx == 12:
            return C_HEADER_RT2_25
        if col_idx == BRAND_COL_IDX:
            return C_HEADER_BRAND
        if col_idx == N_COLS - 1:
            return C_HEADER_RT2_26
        return C_HEADER_26

    def cell_bg(col_idx, row_type):
        if row_type != "normal":
            return ROW_BG[row_type]
        if col_idx == 12:
            return C_CELL_RT2_25
        if col_idx == BRAND_COL_IDX:
            return C_CELL_BRAND
        if col_idx == N_COLS - 1:
            return C_CELL_RT2_26
        return "white"

    rects, colors = [], []
    for col_idx in range(N_COLS):
        rects.append(Rectangle((x_edges[col_idx], 0), COL_WIDTHS[col_idx], row_h))
        colors.append(header_bg(col_idx))
    for r, (vals_25, rt2_25, brand, vals_26, rt2_26, row_type) in enumerate(row_cells, start=1):
        for col_idx in range(N_COLS):
            rects.append(Rectangle((x_edges[col_idx], r), COL_WIDTHS[col_idx], row_h))
            colors.append(cell_bg(col_idx, row_type))
    ax.add_collection(PatchCollection(rects, facecolors=colors, edgecolors=C_GRID, linewidths=0.6))

    # ── Teks sel ──
    def cell_text(col_idx, row_idx, text, fontcolor="black", bold=False, fontsize=9.5):
        x0 = x_edges[col_idx]
        w = COL_WIDTHS[col_idx]
        y0 = row_idx * row_h
        ax.text(x0 + w / 2, y0 + row_h / 2, text, ha="center", va="center",
                fontsize=fontsize, color=fontcolor, fontweight="bold" if bold else "normal")

    for col_idx, label in enumerate(COL_LABELS):
        fontcolor = "white" if col_idx in (12, N_COLS - 1) else "black"
        cell_text(col_idx, 0, label, fontcolor, bold=True, fontsize=8)

    for r, (vals_25, rt2_25, brand, vals_26, rt2_26, row_type) in enumerate(row_cells, start=1):
        fontcolor = "white" if row_type == "divab1" else "black"
        for c, v in enumerate(vals_25):
            cell_text(c, r, v, fontcolor=fontcolor, fontsize=7.5)
        cell_text(12, r, rt2_25, fontcolor=fontcolor, fontsize=7.5)
        cell_text(BRAND_COL_IDX, r, brand, fontcolor=fontcolor, bold=True, fontsize=8)
        for i, v in enumerate(vals_26):
            cell_text(BRAND_COL_IDX + 1 + i, r, v, fontcolor=fontcolor, fontsize=7.5)
        cell_text(N_COLS - 1, r, rt2_26, fontcolor=fontcolor, fontsize=7.5)

    # ── Footer: alamat ──
    fy = grid_bottom_y + footer_text_h * 0.6
    for label in [info["Propinsi"], info["Kota"], info["Kecamatan"], info["Alamat"]]:
        ax.text(cx, fy, label, ha="center", va="center", fontsize=9)
        fy += footer_text_h

    # tidak perlu plt.close(fig) -- fig dibuat lewat Figure() langsung, tidak
    # pernah terdaftar di registry global pyplot, jadi tidak ada yang perlu
    # dibersihkan di sana. Cukup dibiarkan di-garbage-collect begitu fig keluar scope.
    return fig


def _save_fig_bytes(fig: Figure) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200, bbox_inches="tight", pad_inches=0.2, facecolor="white")
    return buf.getvalue()


def _write_bytes_atomic(out_path: Path, data: bytes) -> None:
    # Tulis ke file sementara dulu, baru os.replace() ke nama final (atomic) --
    # kalau langsung tulis ke nama final dan ada 2 request nge-render outlet yang
    # sama bersamaan (mis. 2 tab browser), tulisan byte dari keduanya bisa kebentur
    # di file yang sama dan hasilnya PNG korup/garbled ("burik", nimpa dobel).
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_name(f".{out_path.stem}.{uuid.uuid4().hex}.tmp.png")
    try:
        tmp_path.write_bytes(data)
        # os.replace() bisa gagal SESAAT dengan PermissionError (WinError 32,
        # "process cannot access the file") kalau file .tmp yang baru saja
        # ditulis masih sempat dikunci sebentar oleh proses lain -- paling
        # sering Windows Defender/antivirus real-time scan yang otomatis scan
        # file yang baru dibuat (persis pola yang sudah ditemukan & diperbaiki
        # di omset_pipeline/transpose.py._replace_with_retry() -- terbukti
        # nyata dari error "Copy/Print tidak tersedia: [WinError 32]" yang
        # dilaporkan user). Lock semacam ini biasanya lepas dalam <1-2 detik.
        for attempt in range(6):
            try:
                os.replace(tmp_path, out_path)
                return
            except PermissionError:
                if attempt == 5:
                    raise
                time.sleep(0.5)
    finally:
        tmp_path.unlink(missing_ok=True)


def render_outlet_report(
    site: str | None = None,
    omshar_type: str = "UMUM",
    out_path: Path | None = None,
    with_keg: bool = False,
    *,
    precomputed: tuple | None = None,
) -> tuple[Path, bytes]:
    """Render PNG laporan outlet. Kembalikan (path_file, png_bytes) -- bytes dipakai
    langsung oleh web app (tanpa baca ulang file dari disk), file tetap ditulis ke
    output/IMAGE/ seperti sebelumnya.

    `precomputed` = tuple (row_cells, info, cutoff) dari build_report_rows() yang sudah
    dihitung pemanggil -- dipakai omset_search_app.py supaya query data tidak diulang 3x
    per tampilan (tabel + Excel + PNG). Kalau None, hitung sendiri (jalur CLI)."""
    if not MATPLOTLIB_AVAILABLE:
        raise RuntimeError(
            "matplotlib tidak bisa dimuat -- generate PNG tidak tersedia. Kemungkinan besar "
            "diblokir Windows Smart App Control (cek Settings > Privacy & security > "
            "Windows Security > App & browser control). Pencarian & tabel tetap bisa dipakai "
            f"seperti biasa. Detail error asli: {_MATPLOTLIB_IMPORT_ERROR}"
        )
    if precomputed is None:
        row_cells, info, cutoff = build_report_rows(site, omshar_type, with_keg)
    else:
        row_cells, info, cutoff = precomputed

    png_bytes = _save_fig_bytes(_build_figure(row_cells, info, cutoff))

    if out_path is None:
        safe_name = "".join(c for c in f"{info['Site']} {info['Outlet']}" if c not in '\\/:*?"<>|')
        out_path = OUTPUT_DIR / f"{safe_name}.png"
    _write_bytes_atomic(out_path, png_bytes)

    return out_path, png_bytes


if __name__ == "__main__":
    site = input("Site number: ").strip()
    group = input("Grup [UMUM/HOREKA, default UMUM]: ").strip().upper() or "UMUM"
    path, _ = render_outlet_report(site, group)
    print(f"Tersimpan: {path}")
