"""
MCLUB PIPELINE
Otomatisasi klasifikasi tier outlet (Gold/Platinum/MCLUB) dari file RAW yang
diterima dari divisi lain (D:\\OUTLET MCLUB PLATINUM GOLD\\RAW\\), menggantikan
proses manual maintain file kerja 55MB/13MB + REKAP + PDF.

Semua rumus di bawah sudah diverifikasi terhadap baris data asli (bukan tebakan)
lewat pembongkaran file RAW vs file kerja yang sudah ada -- lihat memory
project-mclub-tier-investigation untuk detail cara verifikasinya.

Konsep kunci:
  - File RAW cuma bawa ~9 bulan terakhir + Strata/Lapisan yang SUDAH ditentukan
    pihak lain (bukan kita hitung) + data omset brand kompetitor per outlet.
  - Archive (CSV lokal di sini) menyimpan riwayat SEMUA bulan yang pernah masuk,
    di-upsert tiap ada RAW baru -- bulan baru ditambah, bulan yang tumpang tindih
    (biasanya 1-2 bulan terakhir yang masih berjalan) ditimpa versi terbaru.
  - Strata/Lapisan/RT2_25 murni titipan dari RAW (pass-through), TIDAK kita
    hitung ulang -- RT2 25 sendiri tidak bisa direkonstruksi dari 6 bulan yang
    ada di RAW (perlu Jan-Jun 2025 yang tidak pernah kita lihat).
"""

import os
import re
import uuid
from pathlib import Path

import openpyxl
import pandas as pd

ARCHIVE_DIR = Path(r"D:\SDAAREA\mclub_pipeline")
ARCHIVE_PATH = {"UMUM": ARCHIVE_DIR / "archive_umum.csv", "HOREKA": ARCHIVE_DIR / "archive_horeka.csv"}

WORKING_FILE = {
    "UMUM": Path(r"D:\OUTLET MCLUB PLATINUM GOLD\List Outlet G P Mc Umum.xlsx"),
    "HOREKA": Path(r"D:\OUTLET MCLUB PLATINUM GOLD\List Outlet G P Mc Horeka.xlsx"),
}

# Semester yang jadi acuan SMT2/SMT1 saat ini -- sama seperti file kerja yang
# ada sekarang (SMT2 25_BIR = Jul-Des 2025, SMT1 26_BIR = Jan-Jun 2026). Ini
# hardcode sesuai siklus berjalan, sama seperti file Excel aslinya juga fixed
# -- perlu diupdate manual kalau siklus semester berikutnya dimulai.
SMT2_25_MONTHS = ["JUL25", "AGS25", "SEP25", "OKT25", "NOV25", "DES25"]
SMT1_26_MONTHS = ["JAN26", "FEB26", "MAR26", "APR26", "MEI26", "JUN26"]

# File kerja (lama) campur format nama kolom bulan -- sebagian pakai spasi +
# singkatan Inggris ("BIR_JUL 25", "BIR_AUG 25", "BIR_OCT 25"), sebagian tanpa
# spasi + singkatan Indonesia ("BIR_DES25", "BIR_JAN26"). Regex ini terima
# keduanya (spasi opsional), lalu dinormalisasi ke singkatan Indonesia biar
# kunci bulan di archive konsisten satu skema saja.
_BIR_MONTH_RE = re.compile(r"^BIR_([A-Za-z]+)\s?(\d{2})$")
_MONTH_NORMALIZE = {
    "JAN": "JAN", "FEB": "FEB", "MAR": "MAR", "APR": "APR", "MEI": "MEI", "MAY": "MEI",
    "JUN": "JUN", "JUL": "JUL", "AGS": "AGS", "AUG": "AGS", "SEP": "SEP",
    "OKT": "OKT", "OCT": "OKT", "NOV": "NOV", "DES": "DES", "DEC": "DES",
}


def _normalize_month_key(mon: str, yy: str) -> str | None:
    key = _MONTH_NORMALIZE.get(mon.upper())
    return f"{key}{yy}" if key else None


def _atomic_write_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        df.to_csv(tmp, index=False, encoding="utf-8-sig")
        os.replace(tmp, path)
    finally:
        tmp.unlink(missing_ok=True)


# ── PARSE RAW ─────────────────────────────────────────────────────────────────

def parse_raw_umum(path) -> pd.DataFrame:
    """Baca file RAW UMUM (sheet 'DB', header satu baris)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["DB"]
    rows = ws.iter_rows(values_only=True)

    header = None
    for row in rows:
        if row and row[0] == "Wil":
            header = row
            break
    if header is None:
        raise ValueError("Header 'Wil' tidak ditemukan di sheet DB -- format file berubah?")
    col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    month_cols: dict[str, int] = {}
    for name, idx in col.items():
        m = _BIR_MONTH_RE.match(name)
        if m and name != "BIR_RT25":
            mkey = _normalize_month_key(*m.groups())
            if mkey:
                month_cols[mkey] = idx

    # Kolom brand kompetitor: semuanya SETELAH "Help Lapisan" (+ spacer kosong),
    # SEBELUM "Total Musuh" -- deteksi dinamis, bukan hardcode daftar brand,
    # supaya tahan kalau daftar brand berubah antar siklus.
    lapisan_idx = col.get("Help Lapisan")
    total_musuh_idx = col.get("Total Musuh")
    brand_cols: dict[str, int] = {}
    if lapisan_idx is not None:
        end = total_musuh_idx if total_musuh_idx is not None else len(header)
        for i in range(lapisan_idx + 1, end):
            name = header[i]
            if name:
                brand_cols[str(name).strip()] = i

    records = []
    for row in rows:
        site = row[col["cust"]]
        if not site:
            continue
        rec = {
            "Site": str(site).strip(),
            "Wil": row[col.get("Wil")],
            "Depo": row[col.get("Depo")],
            "Cust": row[col.get("nama cust")],
            "Strata": row[col.get("Strata Jln AE")],
            "Lapisan": row[col.get("Help Lapisan")],
            "RT2_25": row[col.get("BIR_RT25")],
        }
        for mkey, idx in month_cols.items():
            rec[mkey] = row[idx] if row[idx] is not None else 0
        for brand, idx in brand_cols.items():
            rec[f"MUSUH_{brand}"] = row[idx] if row[idx] is not None else 0
        records.append(rec)
    return pd.DataFrame(records)


def parse_raw_horeka(path, sheet_name: str = "list") -> pd.DataFrame:
    """Baca file HOREKA (RAW: sheet 'list', file kerja lama: sheet 'LIST' --
    parameter sheet_name karena keduanya pakai layout 2-baris header yang sama
    persis, cuma file kerja punya kolom TAMBAHAN yang sudah dihitung (SMT1 26,
    OMS LOSS, dst) yang secara sengaja TIDAK diambil di sini, biar tidak
    tercampur dengan hasil hitungan kita sendiri di compute_metrics()).

    HOREKA punya satu perbedaan penting dari UMUM: SMT2 25 di sini SUDAH berupa
    kolom pass-through ("Rt2 SM2 25") dari sumbernya, BUKAN dihitung dari 6
    bulan individual (Jul-Des 2025) seperti UMUM -- karena bulan-bulan individual
    2025 itu memang tidak pernah ada di data HOREKA, cuma Des 25 + rata-ratanya
    langsung. Makanya SMT2_25 di sini diisi manual, bukan dihitung compute_metrics.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(all_rows):
        if row and len(row) > 2 and row[2] == "Wil":
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Header 'Wil' tidak ditemukan di sheet list -- format file berubah?")
    row_a, row_b = all_rows[header_idx], all_rows[header_idx + 1]

    col: dict[str, int] = {}
    for i in range(max(len(row_a), len(row_b))):
        label_b = row_b[i] if i < len(row_b) else None
        label_a = row_a[i] if i < len(row_a) else None
        label = label_b if label_b is not None else label_a
        if label is not None:
            col[str(label).strip()] = i

    month_map = {
        "Des 25": "DES25", "Jan 26": "JAN26", "Feb 26": "FEB26", "Mar 26": "MAR26",
        "Apr 26": "APR26", "Mei 26": "MEI26", "Jun 26": "JUN26", "Jul 26": "JUL26",
    }
    month_cols = {v: col[k] for k, v in month_map.items() if k in col}
    rt25_idx = col.get("Rt2 25")
    smt2_idx = col.get("Rt2 SM2 25")

    musuh_start = None
    for i in range(len(row_a)):
        if row_a[i] == "Oms Musuh":
            musuh_start = i
            break
    brand_cols: dict[str, int] = {}
    if musuh_start is not None:
        for i in range(musuh_start, len(row_b)):
            name = row_b[i]
            if name:
                brand_cols[str(name).strip()] = i

    records = []
    for row in all_rows[header_idx + 2:]:
        if not row or len(row) <= col.get("Kode Customer", 0):
            continue
        site = row[col["Kode Customer"]]
        if not site:
            continue
        rec = {
            "Site": str(site).strip(),
            "Wil": row[col.get("Wil")],
            "Grup": row[col.get("Grup")],
            "Cust": row[col.get("Cust")],
            "Lapisan": row[col.get("Flag")],
            "RT2_25": row[rt25_idx] if rt25_idx is not None else None,
        }
        if smt2_idx is not None:
            rec["SMT2_25"] = row[smt2_idx]
        for mkey, idx in month_cols.items():
            rec[mkey] = row[idx] if row[idx] is not None else 0
        for brand, idx in brand_cols.items():
            rec[f"MUSUH_{brand}"] = row[idx] if row[idx] is not None else 0
        records.append(rec)
    return pd.DataFrame(records)


# ── ARCHIVE (upsert histori) ──────────────────────────────────────────────────

def load_archive(category: str) -> pd.DataFrame:
    path = ARCHIVE_PATH[category]
    if path.exists():
        return pd.read_csv(path, dtype={"Site": str}, encoding="utf-8-sig")
    return pd.DataFrame(columns=["Site"])


def save_archive(df: pd.DataFrame, category: str) -> None:
    _atomic_write_csv(df, ARCHIVE_PATH[category])


def merge_into_archive(archive: pd.DataFrame, new_data: pd.DataFrame) -> pd.DataFrame:
    """Upsert: kolom/baris baru ditambahkan, sel yang new_data punya nilai
    (non-null) MENIMPA archive -- ini yang membuat bulan lama tetap tersimpan
    (archive punya, new_data tidak -> archive dipertahankan) sekaligus bulan
    yang direvisi ulang di RAW terbaru (keduanya punya -> new_data menang)."""
    if archive.empty or "Site" not in archive.columns:
        return new_data.copy()

    a = archive.set_index("Site")
    n = new_data.set_index("Site")
    all_cols = list(dict.fromkeys(list(a.columns) + list(n.columns)))
    combined_index = a.index.union(n.index)
    a = a.reindex(index=combined_index, columns=all_cols)
    n = n.reindex(index=combined_index, columns=all_cols)

    result = a.copy()
    for c in all_cols:
        mask = n[c].notna()
        result.loc[mask, c] = n.loc[mask, c]
    return result.reset_index()


# ── SEED ARCHIVE DARI FILE KERJA LAMA (sekali jalan) ─────────────────────────

def seed_archive_from_working_file(category: str) -> pd.DataFrame:
    """Import histori dari file kerja 55MB/13MB yang SUDAH ada, supaya bulan-
    bulan lama (sebelum kita mulai pakai pipeline ini) tidak hilang. Cukup
    dijalankan SEKALI di awal (atau kapan saja mau resync manual). Dispatcher --
    UMUM dan HOREKA pakai layout sheet yang beda sama sekali (lihat masing-
    masing fungsi di bawah)."""
    if category == "HOREKA":
        return _seed_archive_working_horeka()
    return _seed_archive_working_umum()


def _seed_archive_working_horeka() -> pd.DataFrame:
    """File kerja HOREKA pakai sheet 'LIST' dengan layout 2-baris header SAMA
    PERSIS dengan file RAW-nya (cuma lebih banyak kolom hasil hitungan) --
    reuse parse_raw_horeka() langsung, kolom tambahan (SMT1 26, OMS LOSS, dst)
    otomatis diabaikan karena tidak dicari di sana."""
    return parse_raw_horeka(WORKING_FILE["HOREKA"], sheet_name="LIST")


def _seed_archive_working_umum() -> pd.DataFrame:
    path = WORKING_FILE["UMUM"]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["DB"]
    rows = ws.iter_rows(values_only=True)

    header = None
    for row in rows:
        if row and (row[0] == "Wil" or (len(row) > 1 and row[1] == "Wil")):
            header = row
            break
    if header is None:
        raise ValueError("Header tidak ditemukan di file kerja -- cek struktur sheet DB")
    col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    month_cols: dict[str, int] = {}
    for name, idx in col.items():
        m = _BIR_MONTH_RE.match(name)
        if m and "SMT" not in name.upper() and not name.upper().endswith("RT25"):
            mkey = _normalize_month_key(*m.groups())
            if mkey:
                month_cols[mkey] = idx
    rt25_idx = col.get("BIR_RT25")

    site_col = col.get("cust")
    name_col = col.get("nama cust")
    wil_col = col.get("Wil")
    depo_col = col.get("Depo")
    strata_col = col.get("Strata Jln AE")
    lapisan_col = col.get("Help Lapisan")

    records = []
    for row in rows:
        if site_col is None or not row[site_col]:
            continue
        rec = {
            "Site": str(row[site_col]).strip(),
            "Wil": row[wil_col] if wil_col is not None else None,
            "Depo": row[depo_col] if depo_col is not None else None,
            "Cust": row[name_col] if name_col is not None else None,
            "Strata": row[strata_col] if strata_col is not None else None,
            "Lapisan": row[lapisan_col] if lapisan_col is not None else None,
            "RT2_25": row[rt25_idx] if rt25_idx is not None else None,
        }
        for mkey, idx in month_cols.items():
            rec[mkey] = row[idx] if row[idx] is not None else 0
        records.append(rec)
    return pd.DataFrame(records)


# ── HITUNG METRIK (rumus terverifikasi) ──────────────────────────────────────

def _safe_div(a: pd.Series, b: pd.Series) -> pd.Series:
    out = a / b
    return out.replace([float("inf"), float("-inf")], 0).fillna(0)


def compute_metrics(df: pd.DataFrame, category: str = "UMUM") -> pd.DataFrame:
    """category menentukan basis INDUSTRI -- terbukti BEDA antar channel dari
    reverse-engineering nyata (bukan tebakan): UMUM pakai SMT2_25 (semester
    tertutup terakhir) sebagai basis "omset kita" di rumus INDUSTRI, HOREKA
    malah pakai SMT1_26 (semester berjalan). Diverifikasi ke baris asli:
    - UMUM (TOMMY RIYADI): INDUSTRI=47.444 = SMT2_25(22.444) + Musuh(25) exact.
    - HOREKA (GRAND MERCURE): INDUSTRI=13.5 = SMT1_26(0.5) + Musuh(13) exact --
      SMT2_25(1) + Musuh(13) = 14, TIDAK cocok, jadi ini bukan salah baca."""
    df = df.copy()
    present_sm2 = [m for m in SMT2_25_MONTHS if m in df.columns]
    present_sm1 = [m for m in SMT1_26_MONTHS if m in df.columns]

    if len(present_sm2) == len(SMT2_25_MONTHS):
        # SEMUA 6 bulan individual Jul-Des 2025 ada (kasus UMUM) -- itu paling
        # otoritatif, hitung ulang dari situ.
        df["SMT2_25"] = df[present_sm2].fillna(0).mean(axis=1)
    elif "SMT2_25" in df.columns:
        # Tidak lengkap 6 bulan (HOREKA cuma punya DES25 sendiri, bukan 6
        # bulan penuh) -- pakai kolom pass-through ("Rt2 SM2 25") apa adanya,
        # JANGAN dihitung ulang dari bulan yang tidak lengkap (itu bug lama:
        # 1 dari 6 bulan kebaca "present" lalu nimpa pass-through dengan 0).
        df["SMT2_25"] = df["SMT2_25"].fillna(0)
    elif present_sm2:
        # Ada sebagian bulan tapi tidak lengkap, dan tidak ada pass-through --
        # fallback rata-rata dari yang ada (lebih baik dari 0 sama sekali).
        df["SMT2_25"] = df[present_sm2].fillna(0).mean(axis=1)
    else:
        df["SMT2_25"] = 0.0
    df["SMT1_26"] = df[present_sm1].fillna(0).mean(axis=1) if present_sm1 else 0.0
    df["OMS_LOSS"] = df["SMT2_25"] - df["SMT1_26"]
    df["PCT_SMT1_VS_SMT2"] = _safe_div(df["SMT1_26"], df["SMT2_25"])
    df["PCT_SMT1_VS_RT25"] = _safe_div(df["SMT1_26"], df["RT2_25"]) if "RT2_25" in df.columns else 0.0

    musuh_cols = [c for c in df.columns if c.startswith("MUSUH_")]
    df["TOTAL_MUSUH"] = df[musuh_cols].fillna(0).sum(axis=1) if musuh_cols else 0.0
    industri_base = df["SMT1_26"] if category == "HOREKA" else df["SMT2_25"]
    df["INDUSTRI"] = industri_base + df["TOTAL_MUSUH"]
    df["PCT_BIR_VS_MUSUH"] = _safe_div(df["SMT1_26"], df["TOTAL_MUSUH"])
    df["PCT_BIR_VS_INDUSTRI"] = _safe_div(df["SMT1_26"], df["INDUSTRI"])
    return df
